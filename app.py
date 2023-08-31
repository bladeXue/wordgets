import toga
from toga.style import Pack
from toga.style.pack import COLUMN, ROW, LEFT, CENTER, RIGHT, HIDDEN, VISIBLE, NORMAL, BOLD, TRANSPARENT
import os
import json
import re
import copy
from functools import partial
import sqlite3
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests
from supermemo2 import SMTwo
from datetime import datetime
import random
from typing import Union
import time
from threading import Thread
import urllib.parse
import hashlib
import shutil
import asyncio
import urllib.request
import toga.platform 
import zipfile

# Global variables
g_strCurrentOS      = toga.platform.current_platform
g_strRunPath        = ""
g_strDataPath       = ""
g_strLangPath       = ""
g_strDBPath         = ""
g_strConfigPath     = ""
g_strCardsPath      = ""
g_strWordlistsPath  = ""
g_dictCards         = {}
g_dictWordLists     = {}
g_EmptyHtml         = '<!DOCTYPE html> <html> <head> <title> </title> </head> <body> </body> </html>' #Cannot be "". Otherwise, it will not work
g_bIsMobile         = True
g_strAppId          = ""
g_strAccessToken    = ""
g_strUsername       = ""

#Config
g_strDefaultLang        = "English"
g_strOpenedWordListName = ""
g_bOnlyReview           = False

# Color scheme
g_clrBg                     = "#EEF4F9"
g_clrPressedNavigationBtn   = "#E6EBF0"
g_clrStrange                = "#CFBAF0"
g_clrVague                  = "#FDE4CF"
g_clrFamiliar               = "#B9FBC0"
g_clrDelete                 = "#8EECF5"
g_clrOpenedTabText          = "#1651AA"


# Keywords for text display
g_strFront      = ""
g_strBack       = ""
g_strOnesided   = ""
g_strFrontend   = ""
g_strBackend    = ""
g_strFilePath   = ""
g_strCard1      = ""
g_strCard2      = ""

# Widgets texts
g_strLblNewWordsPerGroup                = ""
g_strLblOldWordsPerGroup                = ""
g_strLblReviewPolicy                    = ""
g_strCboReviewPolicy_item_LearnFirst    = ""
g_strCboReviewPolicy_item_Random        = ""
g_strCboReviewPolicy_item_ReviewFirst   = ""

# Messagebox
g_strErrorDialogTitle                       = ""
g_strErrorDialogMsg_EmptyName               = ""
g_strErrorDialogMsg_RepetitiveName          = ""
g_strErrorDialogMsg_InvalidFileName         = ""
g_strErrorDialogMsg_AssociatedWordList      = ""
g_strQuestionDialogTitle                    = ""
g_strQuestionDialogMsg_Delete               = ""
g_strErrorDialogMsg_InvalidFile             = ""
g_strErrorDialogMsg_NoWordList              = ""
g_strErrorDialogMsg_Card1IsNull             = ""
g_strErrorDialogMsg_Card1SameAsCard2        = ""
g_strQuestionDialogMsg_NotFoundWordList     = ""
g_strInfoDialogTitle                        = ""
g_strInfoDialogMsg_Complete                 = ""
g_strErrorDialogMsg_DownloadFailed          = ""
g_strInfoDialogMsg_MissionComplete          = ""
g_strInfoDialogMsg_ContinueToStudyNewWords  = ""
g_strQuestionDialogMsg_NoFurtherReview      = ""
g_strInfoDialogMsg_LoginSuccess             = ""
g_strQuestionDialogMsg_Logout               = ""
g_strInfoDialogMsg_LoginFailure             = ""
g_strQuestionDialogMsg_Disconnected         = ""
g_strInfoDialogMsg_NotSyncYet               = ""
g_strErrorDialogMsg_IncompatibleDB          = ""
g_strQuestionDialogMsg_NullDLink            = "" 
g_strQuestionDialogMsg_BrokenCloudDB        = ""
g_strInfoDialogMsg_AboutAppId               = ""
g_strInfoDialogMsg_SyncCompleted            = ""
g_strErrorDialogMsg_VerificationFailed      = ""
g_strInfoDialogTitle_Attention              = ""
g_strInfoDialogMsg_DoNotClose               = ""
g_strErrorDialogMsg_CanNotMigrate           = ""
g_strLblMigrationInfo_Dyn_Step1             = ""
g_strLblMigrationInfo_Dyn_Step2             = ""
g_strLblMigrationInfo_Dyn_Step3             = ""
g_strLblMigrationInfo_Dyn_Step4             = ""
g_strLblMigrationInfo_Dyn_Step5             = ""
g_strLblMigrationInfo_Dyn_Step6             = ""
g_strInfoDialogMsg_MigrationCompleted       = ""
g_strErrorDialogMsg_MigrationFailed         = ""

#Temporary global
g_strTmpNameTxt             = ""
g_bTmpEditNewCard           = True
g_bTmpEditNewWordList       = True
g_iTmpPrevWordNo            = -1
g_iTmpPrevCardType          = -1
g_bSwitchContinueNewWords   = True
g_iTmpCntStudiedNewWords    = 0
g_iTmpCntStudiedOldWords    = 0 # It can be over than real total of old words.
g_iTmpSyncState             = -2
g_bTmpProgramInitEnd           = False

# Functions
def GetRenderedHTML(strHTMLTemplateFilePath: str, strPyFilePath: str, dicDynVariables: str) -> str:
    folder_path, file_name = os.path.split(strHTMLTemplateFilePath)
    dicDynVariables['folder_path'] = folder_path
    dicDynVariables['file_name'] = file_name
    with open(strPyFilePath, "r",encoding="utf-8") as ofs:
        exec(ofs.read(), globals(),dicDynVariables)
    strRenderedHTML = dicDynVariables.get("output")
    return strRenderedHTML

def ConvertNumToExcelColTitle(n: int) -> str: #Use for convert a number to the Excel column serial name. Pay attention: from 1
    column_title = ""
    while n > 0:
        remainder = (n - 1) % 26
        column_title = chr(65 + remainder) + column_title
        n = (n - 1) // 26
    return column_title

def FilterForExcelNoneValue(strInput: Union[str, None]) -> str:
    if strInput == None:
        return ""
    return strInput

def GetFolderSize(strFolder: str) -> int:
    total_size = 0
    for path, dirs, files in os.walk(strFolder):
        for file in files:
            file_path = os.path.join(path, file)
            total_size += os.path.getsize(file_path)
    return total_size

#Autoplay 1: Enable on Linux p.s. ._impl.native.get_settings().set_media_playback_requires_user_gesture(False) not work
class AutoPlayThread(Thread):  
    def __init__(self, strHtmlText):
        super(AutoPlayThread, self).__init__()
        self._strHtmlText = strHtmlText
    
    def run(self):
        try:
            soup = BeautifulSoup(self._strHtmlText, 'html.parser')
            lsAudioTags = soup.find_all('audio')
            for eachAudioTag in lsAudioTags:
                if eachAudioTag.has_attr('autoplay'):
                    src = eachAudioTag['src']
                    filedir = ""
                    if src.startswith("http://") or src.startswith("https://"):
                        #Web sound
                        response = requests.get(src, stream = True)
                        
                        with open(os.path.join(g_strDataPath,"TmpMusic.snd_"), "wb") as f:
                            for chunk in response.iter_content(chunk_size=1024*1024):
                                if chunk:
                                    f.write(chunk)
                        filedir = os.path.join(g_strDataPath,"TmpMusic.snd_")
                    else:
                        filedir = src
                    from pydub import AudioSegment
                    from pydub.playback import play
                    sound = AudioSegment.from_file(filedir)  #Auto detect format
                    play(sound)
        except Exception as e:
            return
    def kill(self):
        try:
            self._stop()
        except Exception:
            return

autoPlayThread = AutoPlayThread("")
autoPlayThread.setDaemon(True)

# MacOS does not support set_content method. Use flask
class FlaskThreadOnMac(Thread):   
    def __init__(self, strHtmlPath, strPyPath, dicDynVariables):
        super(FlaskThreadOnMac, self).__init__()
        self._strHtmlPath = strHtmlPath
        self._strPyPath = strPyPath
        self._dicDynVariables = dicDynVariables
    
    def run(self):
        GetRenderedHTML(
            self._strHtmlPath,
            self._strPyPath,
            self._dicDynVariables
        )

    def kill(self):
        try:
            self._stop()
        except Exception:
            return

class wordgets(toga.App):
    #Hide menu bar
    def _create_impl(self):
        factory_app = self.factory.App
        factory_app.create_menus = lambda _: None
        return factory_app(interface=self)

    def startup(self):
        if g_strCurrentOS == 'windows':                          # Autoplay 3: Enable on Windows
            from System import Environment
            Environment.SetEnvironmentVariable("WEBVIEW2_ADDITIONAL_BROWSER_ARGUMENTS", "--autoplay-policy=no-user-gesture-required")

        global g_strRunPath, g_strLangPath, g_strDBPath, g_strConfigPath, g_strCardsPath, g_strWordlistsPath, g_strDataPath, g_strAppId, g_strAccessToken, g_strUsername

        g_strRunPath        = self.paths.app.absolute()
        g_strDataPath       = self.paths.data.absolute()

        g_strLangPath       = os.path.join(g_strRunPath, "resources/languages.json")
        g_strDBPath         = os.path.join(g_strDataPath, "wordgets_stat.db")
        g_strConfigPath     = os.path.join(g_strDataPath, "configuration.json")
        g_strCardsPath      = os.path.join(g_strDataPath, "cards.json")
        g_strWordlistsPath  = os.path.join(g_strDataPath, "wordlists.json")
        self.main_window    = toga.MainWindow(title=self.formal_name)
        boxMain             = toga.Box(style = Pack(direction = COLUMN, background_color = g_clrBg))
        boxNavigationBar    = toga.Box(style = Pack(direction = ROW, padding_top = 5, padding_bottom = 5, alignment = CENTER))
        
        self.btnIndex_boxNavigateBar       = toga.Button("",style = Pack(flex = 1), on_press = self.cbBtnIndexOnPress)
        self.btnSettings_boxNavigateBar    = toga.Button("",style = Pack(flex = 1), on_press = self.cbBtnSettingsOnPress)
        self.btnLibrary_boxNavigateBar     = toga.Button("",style = Pack(flex = 1), on_press = self.cbBtnLibraryOnPress)
        
        boxNavigationBar.add(
            self.btnIndex_boxNavigateBar,
            self.btnLibrary_boxNavigateBar,
            self.btnSettings_boxNavigateBar
        )
        self.boxBody             = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        
        boxMain.add(
            boxNavigationBar,
            self.boxBody
        )
        
        # Index body
        self.boxIndexBody = toga.Box(style=Pack(direction = COLUMN, flex = 1))
        boxIndexBody_Row1 = toga.Box(style = Pack(direction = ROW, alignment = CENTER))
        self.cboCurWordList_boxIndexBody = toga.Selection(style = Pack(width = 150))
        self.chkReviewOnly_boxIndexBody = toga.Switch("", style= Pack(background_color = TRANSPARENT), on_change = self.cbChangeChkReviewOnly)
        self.btnSync_boxIndexBody = toga.Button("", style = Pack(width = 150), on_press = self.cbSyncDBToCloud)
        
        boxIndexBody_Row1.add(
            self.cboCurWordList_boxIndexBody,
            self.chkReviewOnly_boxIndexBody,
            self.btnSync_boxIndexBody
        )
        
        boxIndexBody_Row2 = toga.Box(style= Pack(direction = ROW))
        self.lblLearningProgressItem_boxIndexBody = toga.Label("", style = Pack(background_color = TRANSPARENT))
        self.lblLearningProgressValue_boxIndexBody = toga.Label("", style = Pack(background_color = TRANSPARENT))
        boxIndexBody_Row2.add(
            self.lblLearningProgressItem_boxIndexBody,
            self.lblLearningProgressValue_boxIndexBody
        )
        boxIndexBody_Row3 = toga.Box(style= Pack(direction = ROW, alignment = CENTER))

        btnHiddenForAlignment_Row3 = toga.Button("", style = Pack(width = 1, padding_top = 5, padding_bottom = 5, visibility = HIDDEN))
        self.boxIndexBody_Row3Hidden = toga.Box(style=Pack(direction = ROW,alignment = CENTER))
        
        lblDelete_boxIndexBody = toga.Label("🗑️", style = Pack( background_color = TRANSPARENT))
        self.lblDeleteTip_boxIndexBody = toga.Label("",style = Pack( background_color = TRANSPARENT))
        self.boxIndexBody_Row3Hidden.add(
            lblDelete_boxIndexBody,
            self.lblDeleteTip_boxIndexBody
        )
        self.boxIndexBody_Row3Hidden.style.update(visibility = VISIBLE)

        self.boxIndexBody_Row3Blank = toga.Box(style=Pack(direction = COLUMN, flex = 1))
        boxIndexBody_Row3.add(
            btnHiddenForAlignment_Row3,
            self.boxIndexBody_Row3Hidden,
            self.boxIndexBody_Row3Blank
        )
        self.wbWord_boxIndexBody = toga.WebView(style = Pack(direction = COLUMN, flex = 1))

        if g_strCurrentOS == "android":                          # Autoplay 4: Enable on Android
            self.wbWord_boxIndexBody._impl.native.getSettings().setMediaPlaybackRequiresUserGesture(False)
        
        boxIndexBody_Row5Border = toga.Box(style=Pack(direction = ROW))
        boxIndexBody_Row5Hidden = toga.Box(style=Pack(direction = COLUMN))
        btnVoidForRemainRow5_boxIndexBody = toga.Button("",style=Pack(visibility = HIDDEN, width = 1)) #If not, the box will not appear
        boxIndexBody_Row5Hidden.add(
            btnVoidForRemainRow5_boxIndexBody
        )
        boxIndexBody_Row5_outside = toga.Box(style=Pack(direction = COLUMN, flex = 1))
        self.boxIndexBody_Row5 = toga.Box(style=Pack(direction = ROW, flex = 1))# Changing tab will be lost
        boxIndexBody_Row5_outside.add(
            self.boxIndexBody_Row5
        )
        boxIndexBody_Row5Border.add(
            boxIndexBody_Row5Hidden,
            boxIndexBody_Row5_outside
        )

        self.boxIndexBody.add(
            boxIndexBody_Row1,
            boxIndexBody_Row2,
            boxIndexBody_Row3,
            self.wbWord_boxIndexBody,
            boxIndexBody_Row5Border
        )

        # Buttons
        self.btnStrange = toga.Button("", style = Pack(flex = 1, background_color = g_clrStrange))
        self.btnVague   = toga.Button("", style = Pack(flex = 1, background_color = g_clrVague))
        self.btnFamiliar= toga.Button("", style = Pack(flex = 1, background_color = g_clrFamiliar))
        #self.btnDelete  = toga.Button("🗑️", style = Pack(flex = 1, background_color = g_clrDelete))
        self. sldDelete = toga.Slider(style = Pack(flex = 1, background_color = TRANSPARENT, height = 25), value = 0, tick_count = None,
                              on_release= self.cbSldDeleteOnRelease
                              )
        self.btnShowBack= toga.Button("", style = Pack(flex = 1))

        self.btnShowBack.on_press = self.cbShowBack
        self.btnStrange.on_press = partial(self.cbNextCard, 0)
        self.btnVague.on_press = partial(self.cbNextCard, 2)
        self.btnFamiliar.on_press = partial(self.cbNextCard, 4)
        #self.btnDelete.on_press = self.cbNoFurtherReviewThisWord

        # Settngs body
        self.boxSettingsBody = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        boxSettingsBody_Row1 = toga.Box(style = Pack(direction = ROW, alignment = CENTER))

        boxSettingsBody_Row1Col1 = toga.Box(style = Pack(direction = ROW, width = 150))
        self.lblLang_boxSettingsBody        = toga.Label("", style = Pack(background_color = TRANSPARENT))
        lblLangEmoji_boxSettingsBody   = toga.Label("🌐", style = Pack(background_color = TRANSPARENT))

        boxSettingsBody_Row1Col1.add(
            self.lblLang_boxSettingsBody,
            lblLangEmoji_boxSettingsBody
        )

        boxSettingsBody_Row1Col2 = toga.Box(style = Pack(direction = ROW))
        self.cboChangeLang_boxSettingsBody = toga.Selection(style = Pack(width = 300))
        boxSettingsBody_Row1Col2.add(
            self.cboChangeLang_boxSettingsBody
        )

        boxSettingsBody_Row1.add(
            boxSettingsBody_Row1Col1,
            boxSettingsBody_Row1Col2
        )

        boxSettingsBody_Row2                    = toga.Box(style = Pack(direction = ROW, alignment = CENTER, padding_top = 5))
        boxSettingsBody_Row2Col1                = toga.Box(style = Pack(direction = ROW, width = 150))
        self.lblCloudSvc_boxSettingsBody        = toga.Label("", style = Pack(background_color = TRANSPARENT))
        boxSettingsBody_Row2Col1.add(
            self.lblCloudSvc_boxSettingsBody
        )

        boxSettingsBody_Row2Col2 = toga.Box(style = Pack(direction = COLUMN, alignment = CENTER))
        boxSettingsBody_Row2Col2Row1 = toga.Box(style = Pack(direction = ROW, flex = 1, alignment = CENTER))
        self.btnLogin_Row2Col2 = toga.Button("", style = Pack(background_color = TRANSPARENT), on_press = self.cbLoginCloudSvc)
        self.btnLogout_Row2Col2 = toga.Button("", style = Pack(background_color = TRANSPARENT), on_press = self.cbLogoutAccount)
        boxSettingsBody_Row2Col2Row1.add(
            self.btnLogin_Row2Col2,
            self.btnLogout_Row2Col2,
        )

        self.lblLoginState_Row2Col2 = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        boxSettingsBody_Row2Col2.add(
            boxSettingsBody_Row2Col2Row1,
            self.lblLoginState_Row2Col2
        )
        
        boxSettingsBody_Row2.add(
            boxSettingsBody_Row2Col1,
            boxSettingsBody_Row2Col2
        )

        boxSettingsBody_Row3                    = toga.Box(style = Pack(direction = ROW, alignment = CENTER, padding_top = 5))
        boxSettingsBody_Row3Col1                = toga.Box(style = Pack(direction = ROW, width = 150))
        self.lblClearCache_boxSettingsBody      = toga.Label("", style = Pack(background_color = TRANSPARENT))
        boxSettingsBody_Row3Col1.add(
            self.lblClearCache_boxSettingsBody
        )

        boxSettingsBody_Row3Col2                = toga.Box(style = Pack(direction = ROW, alignment = CENTER))
        self.btnClearCache_Row3Col2             = toga.Button("", style = Pack(background_color = TRANSPARENT), on_press = self.cbClearCache)
        self.lblCacheSize_Row3Col2              = toga.Label("")
        boxSettingsBody_Row3Col2.add(
            self.btnClearCache_Row3Col2,
            self.lblCacheSize_Row3Col2
        )
        boxSettingsBody_Row3.add(
            boxSettingsBody_Row3Col1,
            boxSettingsBody_Row3Col2
        )

        boxSettingsBody_Row4                    = toga.Box(style = Pack(direction=ROW, alignment=CENTER, padding_top= 5))
        boxSettingsBody_Row4Col1                = toga.Box(style = Pack(direction = ROW, width = 150))
        self.lblMigration_boxSettingsBody       = toga.Label("", style = Pack(background_color = TRANSPARENT))
        boxSettingsBody_Row4Col1.add(
            self.lblMigration_boxSettingsBody
        )

        boxSettingsBody_Row4Col2                = toga.Box(style = Pack(direction = ROW, alignment = CENTER))
        self.btnMigrateFromAnki_boxSettingsBody    = toga.Button("", style = Pack(background_color = TRANSPARENT), on_press = self.cbMigrateFromAnkiOnPress)
        boxSettingsBody_Row4Col2.add(
            self.btnMigrateFromAnki_boxSettingsBody
        )

        boxSettingsBody_Row4.add(
            boxSettingsBody_Row4Col1,
            boxSettingsBody_Row4Col2
        )


        self.boxSettingsBody.add(
            boxSettingsBody_Row1,
            boxSettingsBody_Row2,
            boxSettingsBody_Row3,
            boxSettingsBody_Row4
        )
        
        # Library body
        self.boxLibraryBody = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        boxLibraryBodyRow1 = toga.Box(style = Pack(direction = ROW, flex = 1))
        boxLibraryBodyRow1Col1 = toga.Box(style = Pack(direction = COLUMN, width = 75))
        self.lblCards_boxLibraryBody = toga.Label("", style = Pack(background_color = TRANSPARENT))
        self.btnAddACard_boxLibraryBody = toga.Button("➕", style = Pack(flex = 1), on_press = self.cbAddACardOnPress)
        boxLibraryBodyRow1Col1.add(
            self.lblCards_boxLibraryBody,
            self.btnAddACard_boxLibraryBody
        )
        sbLibraryBodyRow1Col2 = toga.ScrollContainer(style = Pack(direction = COLUMN, flex = 1, background_color = TRANSPARENT))
        self.boxCards_boxLibraryBody = toga.Box(style = Pack(direction = COLUMN))
        sbLibraryBodyRow1Col2.content = self.boxCards_boxLibraryBody
        boxLibraryBodyRow1.add(
            boxLibraryBodyRow1Col1,
            sbLibraryBodyRow1Col2
        )
        
        boxLibraryBodyRow2 = toga.Box(style = Pack(direction = ROW, flex = 1))
        boxLibraryBodyRow2Col1 = toga.Box(style = Pack(direction = COLUMN, width = 75))
        self.lblWordLists_boxLibraryBody = toga.Label("", style = Pack(background_color = TRANSPARENT))
        self.btnAddAWordList_boxLibraryBody = toga.Button("➕", style = Pack(flex = 1), on_press = self.cbAddAWordListdOnPress)
        boxLibraryBodyRow2Col1.add(
            self.lblWordLists_boxLibraryBody,
            self.btnAddAWordList_boxLibraryBody
        )
        sbLibraryBodyRow2Col2 = toga.ScrollContainer(style = Pack(direction = COLUMN, flex = 1, background_color = TRANSPARENT))
        self.boxWordLists_boxLibraryBody = toga.Box(style = Pack(direction = COLUMN))
        sbLibraryBodyRow2Col2.content = self.boxWordLists_boxLibraryBody
        boxLibraryBodyRow2.add(
            boxLibraryBodyRow2Col1,
            sbLibraryBodyRow2Col2
        )
        self.btnApplyChanges_boxLibraryBody = toga.Button("", on_press = self.cbApplyChangesOnPress)
        self.boxLibraryBody.add(
            self.btnApplyChanges_boxLibraryBody,
            boxLibraryBodyRow1,
            boxLibraryBodyRow2
        )   #button will disappear on android if it is on the bottom

        # Body of editing a card   
        self.boxEditCardBody = toga.Box(style = Pack(direction = COLUMN, flex = 1))

        sbEditCardBodyRow1 = toga.ScrollContainer(style = Pack(direction = COLUMN, flex = 1, background_color = TRANSPARENT), horizontal = False)

        self.lblEditCard_boxEditCardBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1, font_weight = BOLD))
        boxEditCardBodyRow1Row2 = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        self.lblCardName_boxEditCardBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtCardName_boxEditCardBody = toga.TextInput(style = Pack(flex = 1), on_change = self.cbValidateNameTxt)
        boxEditCardBodyRow1Row2.add(
            self.lblCardName_boxEditCardBody,
            self.txtCardName_boxEditCardBody
        )
        boxEditCardBodyRow1Row3 = toga.Box(style = Pack(direction = COLUMN,flex = 1))
        self.lblCardFrontFrontend_boxEditCardBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtCardFrontFrontend_boxEditCardBody = toga.TextInput(style = Pack(flex = 1))
        self.lblCardFrontBackend_boxEditCardBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtCardFrontBackend_boxEditCardBody = toga.TextInput(style = Pack(flex = 1))
        boxEditCardBodyRow1Row3.add(
            self.lblCardFrontFrontend_boxEditCardBody,
            self.txtCardFrontFrontend_boxEditCardBody,
            self.lblCardFrontBackend_boxEditCardBody,
            self.txtCardFrontBackend_boxEditCardBody
        )
        self.chkEnableCardBack = toga.Switch("", style = Pack(background_color = TRANSPARENT, flex = 1))

        self.boxEditCardBodyRow1Row5 = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        self.lblCardBackFrontend_boxEditCardBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtCardBackFrontend_boxEditCardBody = toga.TextInput(style = Pack(flex = 1))
        self.lblCardBackBackend_boxEditCardBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtCardBackBackend_boxEditCardBody = toga.TextInput(style = Pack(flex = 1))
        self.boxEditCardBodyRow1Row5.add(
            self.lblCardBackFrontend_boxEditCardBody,
            self.txtCardBackFrontend_boxEditCardBody,
            self.lblCardBackBackend_boxEditCardBody,
            self.txtCardBackBackend_boxEditCardBody
        )

        sbEditCardBodyRow1.content = toga.Box(
            children=[
                self.lblEditCard_boxEditCardBody,
                boxEditCardBodyRow1Row2,
                boxEditCardBodyRow1Row3,
                self.chkEnableCardBack,
                self.boxEditCardBodyRow1Row5
            ],style = Pack(direction = COLUMN)
        )
        
        self.chkEnableCardBack.on_change = self.cbEnableCardBackOnChange
        self.chkEnableCardBack.value = True  #A bug. Initially not hide

        self.btnSaveCard_boxEditCardBody = toga.Button("", on_press = self.cbSaveCardOnPress)
        self.boxEditCardBody.add(
            sbEditCardBodyRow1,
            self.btnSaveCard_boxEditCardBody
        )

        # Body of editing a word list
        self.boxEditWordListBody = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        sbEditWordListBodyRow1 = toga.ScrollContainer(style = Pack(direction = COLUMN, flex = 1, background_color = TRANSPARENT), horizontal = False)
        self.lblEditWordlist_boxEditWordListBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1, font_weight = BOLD))
        boxEditWordListBodyRow1Row2 = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        self.lblWordListName_boxEditWordListBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtWordListName_boxEditWordListBody = toga.TextInput(style = Pack(flex = 1), on_change = self.cbValidateNameTxt)
        boxEditWordListBodyRow1Row2.add(
            self.lblWordListName_boxEditWordListBody,
            self.txtWordListName_boxEditWordListBody
        )
        boxEditWordListBodyRow1Row3 = toga.Box(style = Pack(direction = COLUMN,flex = 1))
        self.lblWordListXlsx_boxEditWordListBody = toga.Label("", style = Pack(background_color = TRANSPARENT, flex = 1))
        self.txtWordListXlsx_boxEditWordListBody = toga.TextInput(style = Pack(flex = 1))
        
        boxEditWordListBodyRow1Row3.add(
            self.lblWordListXlsx_boxEditWordListBody,
            self.txtWordListXlsx_boxEditWordListBody,
        )

        sbEditWordListBodyRow1.content = toga.Box(
            children=[
                self.lblEditWordlist_boxEditWordListBody,
                boxEditWordListBodyRow1Row2,
                boxEditWordListBodyRow1Row3
            ],style = Pack(direction = COLUMN)
        )

        self.btnSaveWordList_boxEditWordListBody = toga.Button("", on_press = self.cbSaveWordListOnPress)
        self.boxEditWordListBody.add(
            sbEditWordListBodyRow1,
            self.btnSaveWordList_boxEditWordListBody
        )

        # Windows of Cloud Services. Begin------------------------------------------------------------------------------------------------------------------
        ## Procedure 1: Select one cloud storage service provider
        self.boxCloudSvcP1 = toga.Box(style = Pack(direction = COLUMN))
        self.lblCloudSvcP1Title_boxCloudSvcP1 = toga.Label("", style = Pack(font_weight = BOLD)) 
        self.btnCloudSvcP1_BaiduNetdisk_boxCloudSvcP1 = toga.Button("", style = Pack(flex = 1), on_press = self.cbSetCloudSvc)
        self.boxCloudSvcP1.add(
            self.lblCloudSvcP1Title_boxCloudSvcP1,
            self.btnCloudSvcP1_BaiduNetdisk_boxCloudSvcP1
        )

        ## Procedure 2: Enter AppId
        self.boxCloudSvcP2 = toga.Box(style = Pack(direction = COLUMN))
        self.lblCloudSvcP2Title_boxCloudSvcP2 = toga.Label("", style = Pack(font_weight = BOLD))
        self.btnCloudSvcP2WhatIsThis_boxCloudSvcP2 = toga.Button("", on_press = lambda: self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_AboutAppId))
        self.txtCloudSvcP2AppId_boxCloudSvcP2 = toga.TextInput(placeholder="AppKey")
        self.btnCloudSvcP2Next_boxCloudSvcP2 = toga.Button("", on_press = self.cbLoginAccount)
        self.boxCloudSvcP2.add(
            self.lblCloudSvcP2Title_boxCloudSvcP2,
            self.btnCloudSvcP2WhatIsThis_boxCloudSvcP2,
            self.txtCloudSvcP2AppId_boxCloudSvcP2,
            self.btnCloudSvcP2Next_boxCloudSvcP2
        )

        ## Procedure 3: Login account
        self.boxCloudSvcP3 = toga.Box(style = Pack(direction = COLUMN, flex = 1))
        boxCloudSvcP3_row1 = toga.Box(style = Pack(direction = ROW, alignment = CENTER))
        self.lblCloudSvcP3Title_boxCloudSvcP3 = toga.Label("", style = Pack(font_weight = BOLD, flex = 1))
        boxCloudSvcP3_row1.add(
            self.lblCloudSvcP3Title_boxCloudSvcP3
        )

        self.wbCloudSvcP3_LoginAccount = toga.WebView(style = Pack(direction = COLUMN, flex = 1)) #A bug. Cannot automatically adjust height

        boxManualVerification_boxCloudSvcP3 = toga.Box(style = Pack(direction = ROW, alignment = CENTER))
        self.btnManualVerification = toga.Button("", style = Pack(direction = COLUMN, flex = 1), on_press = self.cbManualVerificationOnPress)
        boxManualVerification_boxCloudSvcP3.add(
            self.btnManualVerification
        )
        
        self.boxCloudSvcP3.add(
            boxCloudSvcP3_row1,
            boxManualVerification_boxCloudSvcP3,
            self.wbCloudSvcP3_LoginAccount
        )

        if g_strCurrentOS != 'macOS':
            self.wbCloudSvcP3_LoginAccount.set_content("NO_CONTENT", g_EmptyHtml)
        else:
            self.wbCloudSvcP3_LoginAccount.url = 'http://127.0.0.1'

        # Windows of Cloud Services. End--------------------------------------------------------------------------------------------------------------------

        # Body for migration
        self.boxMigration = toga.Box(style = Pack(direction = COLUMN))
        
        boxRow1_boxMigration = toga.Box(style=Pack(direction=ROW))
        self.lblApkgFile_boxMigration = toga.Label("",style=Pack(width =180, background_color = TRANSPARENT))
        self.txtApkgFilePath_boxMigration = toga.TextInput(value="",style=Pack(flex = 1))
        boxRow1_boxMigration.add(
            self.lblApkgFile_boxMigration,
            self.txtApkgFilePath_boxMigration
        )

        boxRow2_boxMigration = toga.Box(style=Pack(direction=ROW, padding_top = 5))
        self.lblSelectWordList_boxMigration = toga.Label("",style=Pack(width =180, background_color = TRANSPARENT))
        self.cboWordLists_boxMigration = toga.Selection(style=Pack(flex = 1), items = None)
        boxRow2_boxMigration.add(
            self.lblSelectWordList_boxMigration,
            self.cboWordLists_boxMigration
        )

        self.btnMigrate_boxMigration = toga.Button("",style=Pack(flex = 1,padding_top=20),on_press=self.cbMigrate)
        self.lblMigrationInfo_boxMigration = toga.Label("", style=Pack(flex = 1, background_color = TRANSPARENT))
        self.prgRateOfMigration_boxMigration = toga.ProgressBar(style=Pack(flex = 1))

        self.boxMigration.add(
            boxRow1_boxMigration,
            boxRow2_boxMigration,
            self.btnMigrate_boxMigration,
            self.lblMigrationInfo_boxMigration,
            self.prgRateOfMigration_boxMigration
        )

        # Read language file
        self.m_dictLanguages = {}
        with open(g_strLangPath, "r") as f:
            self.m_dictLanguages = json.load(f)

        global g_bIsMobile, g_strDefaultLang, g_strOpenedWordListName, g_bOnlyReview
        if not(g_strCurrentOS == "android" or g_strCurrentOS == "iOS"):
            g_bIsMobile = False

        #Read configuration
        (iX, iY) = self.main_window.position
        (iWidth, iHeight) = self.main_window.size
        if os.path.exists(g_strConfigPath) == True:
            with open(g_strConfigPath, "r") as f:
                dictConfig = json.load(f)
                g_strDefaultLang = dictConfig['lang']
                g_strOpenedWordListName = dictConfig['cur_wordlist']
                g_bOnlyReview = dictConfig['only_review']
                g_strAppId = dictConfig["appid"]
                g_strAccessToken =dictConfig["access_token"]
                g_strUsername =dictConfig["username"]
                if g_bIsMobile == False:
                    t_iX = dictConfig['X']
                    t_iY = dictConfig['Y']
                    t_iWidth = dictConfig['width']
                    t_iHeight = dictConfig['height']
                    #Current lack resolution getter
                    if 0 <= t_iX <= t_iX + t_iWidth  and\
                        0 <= t_iY <= t_iY + t_iHeight :
                        iX = t_iX
                        iY = t_iY
                        iWidth = t_iWidth
                        iHeight = t_iHeight
        
        self.main_window.content = boxMain
        self.main_window.position = (iX, iY)
        self.main_window.size = (iWidth, iHeight)
        
        #Initialize
        self.cbBtnIndexOnPress(None)
        self.main_window.show()

        self.cboChangeLang_boxSettingsBody.items = list(self.m_dictLanguages.keys())
        self.cboChangeLang_boxSettingsBody.value = g_strDefaultLang

        self.cboChangeLang_boxSettingsBody.on_select = self.cbChangeLangOnSelect
        self.ChangeLangAccordingToDefaultLang()

        self.chkReviewOnly_boxIndexBody.value = g_bOnlyReview
        
        if os.path.exists(g_strWordlistsPath) == True:
            with open(g_strWordlistsPath, "r") as f:
                global g_dictWordLists
                g_dictWordLists = json.load(f)
                
        if os.path.exists(g_strCardsPath) == True:
            with open(g_strCardsPath, "r") as f:
                global g_dictCards
                g_dictCards = json.load(f)
        
        if g_strCurrentOS != 'macOS':
            self.wbWord_boxIndexBody.set_content("NO_CONTENT", g_EmptyHtml)
        else:
            self.wbWord_boxIndexBody.url = 'http://127.0.0.1'

        self.m_lsRememberSeq = []
        bAllFilesAreValid = self.ValidateFiles()
        
        self.boxIndexBody_Row3Hidden.style.update(visibility = HIDDEN)

        if g_strUsername != "":
            self.btnSync_boxIndexBody.enabled = True
            # A bug. On Android, async startup will not display anything and asyncio.run will cause confliction and nest_asyncio will raise NotImplementedError
            # Indirect method: use a widget callback to execute async function. This method is not suitable for Windows as it will cause exception!!
            if g_strCurrentOS == 'android':
                chkTmp = toga.Switch("114514", value = False)
                chkTmp.on_change = self.SyncDBToLocal
                chkTmp.value = True
            else:
                asyncio.run(self.SyncDBToLocal(None))
        else:
            self.btnSync_boxIndexBody.enabled = False
        global g_bTmpProgramInitEnd
        g_bTmpProgramInitEnd = True

        if bAllFilesAreValid == True and os.path.exists(g_strDBPath) == True: # Not support displaying an error dialog when no operation on the just executed program. If so, re-config is required.
            lsValidWordlists = []
            conn = sqlite3.connect(g_strDBPath)
            cursor = conn.cursor()
            for eachWordlist in g_dictWordLists:
                cursor.execute("SELECT * FROM statistics WHERE wordlist = ?", (eachWordlist,))
                result = cursor.fetchall()
                if len(result) == 0:
                    continue
                lsValidWordlists.append(eachWordlist)
            conn.commit()
            conn.close()

            if g_strOpenedWordListName not in lsValidWordlists:
                g_strOpenedWordListName = lsValidWordlists[0]
            
            self.cboCurWordList_boxIndexBody.items = lsValidWordlists
            self.cboCurWordList_boxIndexBody.value = g_strOpenedWordListName


            self.cboCurWordList_boxIndexBody.on_select = self.cbChangeCurWordList
            
            self.GenerateNewWordsSeqOfCurrentWordList()
            self.cbNextCard(-1, None)

    # Callback functions
    def cbBtnIndexOnPress(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.btnIndex_boxNavigateBar.style.update(font_weight = BOLD, background_color = g_clrPressedNavigationBtn, color = g_clrOpenedTabText)
        self.btnSettings_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.btnLibrary_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.boxBody.add(
            self.boxIndexBody
        )
        self.SaveConfiguration()

    def cbBtnSettingsOnPress(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.btnIndex_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.btnSettings_boxNavigateBar.style.update(font_weight = BOLD, background_color = g_clrPressedNavigationBtn, color = g_clrOpenedTabText)
        self.btnLibrary_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')

        self.boxBody.add(
            self.boxSettingsBody
        )

        self.lblCacheSize_Row3Col2.text = str(round(GetFolderSize(os.path.join(g_strDataPath, "Cache")) / 1024 / 1024, 2))+ " MB"
        self.ChangeAccountBtnState()
        self.SaveConfiguration()

    def cbBtnLibraryOnPress(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.btnIndex_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.btnSettings_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.btnLibrary_boxNavigateBar.style.update(font_weight = BOLD, background_color = g_clrPressedNavigationBtn, color = g_clrOpenedTabText)
        self.boxBody.add(
            self.boxLibraryBody
        )
        self.SaveConfiguration()
        
        self.m_dictTmpCards = copy.deepcopy(g_dictCards)
        self.m_dictTmpWordLists = copy.deepcopy(g_dictWordLists)
        self.FreshCards()
        self.FreshWordLists()


    def cbChangeLangOnSelect(self, widget):
        global g_strDefaultLang
        g_strDefaultLang = widget.value
        self.ChangeLangAccordingToDefaultLang()
        
    def cbValidateNameTxt(self, widget):
        global g_strTmpNameTxt
        if widget.value == "":
            g_strTmpNameTxt = ""
            return
        if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", widget.value):
            widget.value = g_strTmpNameTxt
        if len(widget.value) > 10:
            widget.value = g_strTmpNameTxt
        g_strTmpNameTxt = widget.value
        
    
    def cbAddACardOnPress(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditCardBody)
        while len(self.boxBody.children) != 0: # A BUG: It must be execute twice at the first time, or the scrollcontainer opened at the first time could be incomplete
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditCardBody)
        self.txtCardName_boxEditCardBody.value = ""
        self.txtCardFrontFrontend_boxEditCardBody.value = ""
        self.txtCardFrontBackend_boxEditCardBody.value = ""
        self.txtCardBackFrontend_boxEditCardBody.value = ""
        self.txtCardBackBackend_boxEditCardBody.value = ""
        global g_bTmpEditNewCard
        g_bTmpEditNewCard = True
        self.txtCardName_boxEditCardBody.readonly = False
    
    def cbAddAWordListdOnPress(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditWordListBody)
        while len(self.boxBody.children) != 0: # A BUG: It must be execute twice at the first time, or the scrollcontainer opened at the first time could be incomplete
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditWordListBody)
        
        self.txtWordListName_boxEditWordListBody.value = ""
        self.txtWordListXlsx_boxEditWordListBody.value = ""
        global g_bTmpEditNewWordList
        g_bTmpEditNewWordList = True
        self.txtWordListName_boxEditWordListBody.readonly = False

    def cbSaveCardOnPress(self, widget):
        strCardCname = self.txtCardName_boxEditCardBody.value
        if strCardCname == "":
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_EmptyName)
            return
        if strCardCname in self.m_dictTmpCards and g_bTmpEditNewCard == True:
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_RepetitiveName)
            return
        
        strFrontFrontendFilePath = self.txtCardFrontFrontend_boxEditCardBody.value
        strFrontBackendFilePath = self.txtCardFrontBackend_boxEditCardBody.value
        strBackFrontendFilePath = self.txtCardBackFrontend_boxEditCardBody.value
        strBackBackendFilePath = self.txtCardBackBackend_boxEditCardBody.value

        bEnableBack = self.chkEnableCardBack.value
        if bEnableBack == False:
            if strFrontFrontendFilePath == "" or strFrontBackendFilePath == "":
                self.main_window.error_dialog(title = g_strErrorDialogTitle, message = g_strErrorDialogMsg_InvalidFileName)
                return
            else:
                if strFrontFrontendFilePath[:4] == "http":
                    try:
                        response = requests.get(strFrontFrontendFilePath, stream = True)
                        filename = strCardCname+"-front.html"
                        strFileWriteTo = os.path.join(g_strDataPath, filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strFrontFrontendFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return
                    
                if strFrontBackendFilePath[:4] == "http":
                    try:
                        response = requests.get(strFrontBackendFilePath, stream = True)
                        filename = strCardCname+"-front.py"
                        strFileWriteTo = os.path.join(g_strDataPath,filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strFrontBackendFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return
                    
                self.m_dictTmpCards[strCardCname] = {
                    "single-sided": {
                        "frontend": strFrontFrontendFilePath,
                        "backend": strFrontBackendFilePath
                    }
                }
        else:
            if strFrontFrontendFilePath == "" or strFrontBackendFilePath == "" or strBackFrontendFilePath == "" or strBackBackendFilePath == "":
                self.main_window.error_dialog(title = g_strErrorDialogTitle, message = g_strErrorDialogMsg_InvalidFileName)
                return
            else:
                if strFrontFrontendFilePath[:4] == "http":
                    try:
                        response = requests.get(strFrontFrontendFilePath, stream = True)
                        filename = strCardCname+"-front.html"
                        strFileWriteTo = os.path.join(g_strDataPath, filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strFrontFrontendFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return
                    
                if strFrontBackendFilePath[:4] == "http":
                    try:
                        response = requests.get(strFrontBackendFilePath, stream = True)
                        filename = strCardCname+"-front.py"
                        strFileWriteTo = os.path.join(g_strDataPath, filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strFrontBackendFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return
                if strBackFrontendFilePath[:4] == "http":
                    try:
                        response = requests.get(strBackFrontendFilePath, stream = True)
                        filename = strCardCname+"-back.html"
                        strFileWriteTo = os.path.join(g_strDataPath, filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strBackFrontendFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return
                    
                if strBackBackendFilePath[:4] == "http":
                    try:
                        response = requests.get(strBackBackendFilePath, stream = True)
                        filename = strCardCname+"-back.py"
                        strFileWriteTo = os.path.join(g_strDataPath, filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strBackBackendFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return

                self.m_dictTmpCards[strCardCname] = {
                    "front": {
                        "frontend": strFrontFrontendFilePath,
                        "backend": strFrontBackendFilePath
                    },
                    "back": {
                        "frontend": strBackFrontendFilePath,
                        "backend": strBackBackendFilePath
                    }
                }
        self.BackToLibraryAndFresh()

    def cbSaveWordListOnPress(self, widget):
        strWordListCname = self.txtWordListName_boxEditWordListBody.value
        if strWordListCname == "":
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_EmptyName)
            return
        if strWordListCname in self.m_dictTmpWordLists and g_bTmpEditNewWordList == True:
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_RepetitiveName)
            return 
        strExcelFilePath = self.txtWordListXlsx_boxEditWordListBody.value

        if strExcelFilePath == "":
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_InvalidFileName)
            return
        else:
            if strExcelFilePath[:4] == "http":
                    try:
                        response = requests.get(strExcelFilePath, stream = True)
                        filename = strWordListCname + ".xlsx"
                        strFileWriteTo = os.path.join(g_strDataPath, filename)
                        with open(strFileWriteTo, "wb") as f:
                            for chunk in response.iter_content(chunk_size = 1024*1024):
                                if chunk:
                                    f.write(chunk)
                        strExcelFilePath = strFileWriteTo
                    except Exception:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                        return
            self.m_dictTmpWordLists[strWordListCname] = {
                "FilePath": strExcelFilePath,
                "CardType1": "",
                "CardType2": "",
                "NewWordsPerGroup": 20,
                "OldWordsPerGroup": 20,
                "policy": 0
            }
        
        self.BackToLibraryAndFresh()
        
    def cbEnableCardBackOnChange(self, widget):
        if widget.value == True:
            self.boxEditCardBodyRow1Row5.style.update(visibility = VISIBLE)
        else:
            self.boxEditCardBodyRow1Row5.style.update(visibility = HIDDEN)

    def cbDeleteCard(self, strCardName, widget):
        setAllUsedCardNames = set()
        for eachWordList in self.m_dictTmpWordLists:
            strWordListCardType1 = self.m_dictTmpWordLists[eachWordList]['CardType1']
            if strWordListCardType1 != "":
                setAllUsedCardNames.add(strWordListCardType1)
            strWordListCardType2 = self.m_dictTmpWordLists[eachWordList]['CardType2']
            if strWordListCardType2 != "":
                setAllUsedCardNames.add(strWordListCardType2) 
        if strCardName in setAllUsedCardNames:
            self.main_window.error_dialog(title = g_strErrorDialogTitle, message = g_strErrorDialogMsg_AssociatedWordList)
            return
        self.m_dictTmpCards.pop(strCardName)
        self.FreshCards()
        self.FreshWordLists()

    def cbEditExistedCard(self, strCardName, widget):
        self.txtCardName_boxEditCardBody.readonly = True
        self.txtCardName_boxEditCardBody.value = strCardName
        if len(self.m_dictTmpCards[strCardName]) == 1:
            self.txtCardFrontFrontend_boxEditCardBody.value = self.m_dictTmpCards[strCardName]['single-sided']['frontend']
            self.txtCardFrontBackend_boxEditCardBody.value = self.m_dictTmpCards[strCardName]['single-sided']['backend']
            self.chkEnableCardBack.value = False
            self.txtCardBackFrontend_boxEditCardBody.value = ""
            self.txtCardBackBackend_boxEditCardBody.value = ""
        else:
            self.txtCardFrontFrontend_boxEditCardBody.value = self.m_dictTmpCards[strCardName]['front']['frontend']
            self.txtCardFrontBackend_boxEditCardBody.value = self.m_dictTmpCards[strCardName]['front']['backend']
            self.chkEnableCardBack.value = True
            self.txtCardBackFrontend_boxEditCardBody.value = self.m_dictTmpCards[strCardName]['back']['frontend']
            self.txtCardBackBackend_boxEditCardBody.value = self.m_dictTmpCards[strCardName]['back']['backend']
        global g_bTmpEditNewCard
        g_bTmpEditNewCard = False
        while len(self.boxBody.children) != 0: #If once, it will not display completely on windows
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditCardBody)
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditCardBody)

    async def cbDeleteWordList(self, widget):
        if await self.main_window.question_dialog(title = g_strQuestionDialogTitle, message = g_strQuestionDialogMsg_Delete):
            self.m_dictTmpWordLists.pop(widget.id)  # An unreasonable solution
            self.FreshCards()
            self.FreshWordLists()
        else:
            return
    
    def cbChangeReviewPolicy(self, strWordList, widget):
        self.m_dictTmpWordLists[strWordList]['policy'] = widget.items.index(widget.value)

    def cbChangeNewWordsPerGroup(self, strWordList, widget):
        try:
            self.m_dictTmpWordLists[strWordList]['NewWordsPerGroup'] = int(widget.value)
        except Exception:
            self.m_dictTmpWordLists[strWordList]['NewWordsPerGroup'] = 1

    def cbChangeOldWordsPerGroup(self, strWordList, widget):
        try:
            self.m_dictTmpWordLists[strWordList]['OldWordsPerGroup'] = int(widget.value)
        except Exception:
            self.m_dictTmpWordLists[strWordList]['OldWordsPerGroup'] = 1
    
    def cbChangeWordListCardType(self, strWordList, iDeleteCardType, widget):
        if iDeleteCardType == 1:
            if widget.value == None:
                self.m_dictTmpWordLists[strWordList]['CardType1'] = ""
            else:
                self.m_dictTmpWordLists[strWordList]['CardType1'] = widget.value
        elif iDeleteCardType == 2:
            if widget.value == None:
                self.m_dictTmpWordLists[strWordList]['CardType2'] = ""
            else:
                self.m_dictTmpWordLists[strWordList]['CardType2'] = widget.value

    def cbEditExistedWordList(self, strWordListName, widget):
        self.txtWordListName_boxEditWordListBody.readonly = True
        self.txtWordListName_boxEditWordListBody.value = strWordListName
        self.txtWordListXlsx_boxEditWordListBody.value = self.m_dictTmpWordLists[strWordListName]['FilePath']
        global g_bTmpEditNewWordList
        g_bTmpEditNewWordList = False
        while len(self.boxBody.children) != 0: #If once, it will not display completely on windows
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditWordListBody)
        while len(self.boxBody.children) != 0: 
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxEditWordListBody)
    
    async def cbApplyChangesOnPress(self, widget):
        try:
            # Validate cards
            for eachCard in self.m_dictTmpCards:
                if len(self.m_dictTmpCards[eachCard]) == 1:
                    strHtmlPath = self.m_dictTmpCards[eachCard]['single-sided']['frontend']
                    with open(strHtmlPath, 'r') as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                    strPyPath = self.m_dictTmpCards[eachCard]['single-sided']['backend']
                    with open(strPyPath, 'r') as f:
                        strCode = f.read()
                        compile(strCode, strPyPath, 'exec')
                else:

                    strFHtmlPath = self.m_dictTmpCards[eachCard]['front']['frontend']
                    with open(strFHtmlPath, 'r') as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                    strFPyPath = self.m_dictTmpCards[eachCard]['front']['backend']
                    with open(strFPyPath, 'r') as f:
                        strCode = f.read()
                        compile(strCode, strFPyPath, 'exec')

                    strBHtmlPath = self.m_dictTmpCards[eachCard]['back']['frontend']
                    with open(strBHtmlPath, 'r') as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                    strBPyPath = self.m_dictTmpCards[eachCard]['back']['backend']
                    with open(strBPyPath, 'r') as f:
                        strCode = f.read()
                        compile(strCode, strBPyPath, 'exec')

            for eachWordlist in self.m_dictTmpWordLists:  
            # Validate wordlists
                strXlsxPath = self.m_dictTmpWordLists[eachWordlist]['FilePath'] 
                xlsx = load_workbook(strXlsxPath, read_only = True)
                sheet = xlsx.worksheets[0]
                xlsx.close()
        except Exception:
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_InvalidFile)
            return
        
        # Check if there is no word list
        if len(self.m_dictTmpWordLists) == 0:
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_NoWordList)
            return
        
        # Check all word list is available
        for eachWordList in self.m_dictTmpWordLists:
            if self.m_dictTmpWordLists[eachWordList]['CardType1'] == "":
                self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_Card1IsNull)
                return
            else:
                if self.m_dictTmpWordLists[eachWordList]['CardType2'] == self.m_dictTmpWordLists[eachWordList]['CardType1']:
                    self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_Card1SameAsCard2)
                    return
        
        conn = sqlite3.connect(g_strDBPath)
        cursor = conn.cursor()

        #If db is not existed,
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS statistics(
                wordlist        TEXT,
                word_no         INTEGER,
                card_type       INTEGER,
                review_date     DATE,
                easiness        REAL,
                interval        INTEGER,
                repetitions     INTEGER
            )
        ''')

        # If word list in db is not existed in current setting, remove
        cursor.execute("SELECT DISTINCT wordlist FROM statistics")
        lsAllWordListsInDB = [elem[0] for elem in cursor.fetchall()]
        lsAllWordListsInDBRemained = lsAllWordListsInDB[:]
        for eachWordListInDB in lsAllWordListsInDB:
            if eachWordListInDB not in self.m_dictTmpWordLists:
                # In case of saving the record
                if await self.main_window.question_dialog(g_strQuestionDialogTitle, message = g_strQuestionDialogMsg_NotFoundWordList % eachWordListInDB):
                    cursor.execute(f"DELETE FROM statistics WHERE wordlist='{eachWordListInDB}'")  
                    lsAllWordListsInDBRemained.remove(eachWordListInDB)
                else:
                    continue
        
        for eachWordList in self.m_dictTmpWordLists:
            xlsx = load_workbook(self.m_dictTmpWordLists[eachWordList]['FilePath'], read_only = True)
            sheet = xlsx.worksheets[0]
            iSheetMaxRow = sheet.max_row
            xlsx.close()
            if eachWordList in lsAllWordListsInDBRemained:
                # If the word list is in db,  detect number change and update card
                cursor.execute(f"SELECT MAX(word_no) FROM statistics WHERE wordlist=?",(eachWordList,)) 
                iWordListMaxRowInDB = cursor.fetchone()[0]
               
                if self.m_dictTmpWordLists[eachWordList]['CardType2'] == "":    # If the word list does not has card2 but database has, delete
                    cursor.execute("DELETE FROM statistics WHERE wordlist=? AND card_type=?",(eachWordList, 2))
                else:                                                   # If the word list has card2 but database does not have, new
                    cursor.execute('SELECT COUNT(*) FROM statistics WHERE wordlist=? AND card_type=?', (eachWordList, 2))
                    iCntWordListCardType2InDB = cursor.fetchone()[0]
                    if iCntWordListCardType2InDB == 0:
                        for i in range(2, iWordListMaxRowInDB + 1):
                            cursor.execute("INSERT INTO statistics (wordlist, word_no, card_type) VALUES (?, ?, ?)", (eachWordList, i, 2))
                
                if iWordListMaxRowInDB < iSheetMaxRow:
                    for i in range(iWordListMaxRowInDB + 1, iSheetMaxRow + 1):
                        cursor.execute("INSERT INTO statistics (wordlist, word_no, card_type) VALUES (?, ?, ?)", (eachWordList, i, 1))
                        if self.m_dictTmpWordLists[eachWordList]['CardType2'] != "":
                            cursor.execute("INSERT INTO statistics (wordlist, word_no, card_type) VALUES (?, ?, ?)", (eachWordList, i, 2))
                elif iWordListMaxRowInDB > iSheetMaxRow:
                    #
                    cursor.execute('DELETE FROM statistics WHERE wordlist=? AND word_no BETWEEN ? AND ?', (eachWordList ,iSheetMaxRow + 1, iWordListMaxRowInDB))
                    
            else:
                # If the word list is not in db, add directly
                for i in range(2, iSheetMaxRow + 1):
                    cursor.execute("INSERT INTO statistics (wordlist, word_no, card_type) VALUES (?, ?, ?)", (eachWordList, i, 1))
                    if self.m_dictTmpWordLists[eachWordList]['CardType2'] != "":
                        cursor.execute("INSERT INTO statistics (wordlist, word_no, card_type) VALUES (?, ?, ?)", (eachWordList, i, 2))
        conn.commit()
        conn.close()
        
        global g_dictWordLists, g_dictCards # Apply 
        g_dictWordLists = copy.deepcopy(self.m_dictTmpWordLists)
        g_dictCards = copy.deepcopy(self.m_dictTmpCards)

        self.m_dictTmpWordLists.clear() #Clear temporary
        self.m_dictTmpCards.clear()
        self.SaveWordListsAndCardsToFiles()

        self.main_window.info_dialog(title = g_strInfoDialogTitle, message = g_strInfoDialogMsg_Complete)

        lsValidWordlists = []
        conn = sqlite3.connect(g_strDBPath)
        cursor = conn.cursor()
        for eachWordlist in g_dictWordLists:
            cursor.execute("SELECT * FROM statistics WHERE wordlist = ?", (eachWordlist,))
            result = cursor.fetchall()
            if len(result) == 0:
                continue
            lsValidWordlists.append(eachWordlist)
        conn.commit()
        conn.close()
        global g_strOpenedWordListName
        if g_strOpenedWordListName not in lsValidWordlists:
            g_strOpenedWordListName = lsValidWordlists[0]
        
        self.cboCurWordList_boxIndexBody.items = lsValidWordlists
        self.cboCurWordList_boxIndexBody.value = g_strOpenedWordListName
        
        if g_strUsername != "":
            await self.SyncDBToLocal(None)

        self.GenerateNewWordsSeqOfCurrentWordList()
        self.cbNextCard(-1, None)

        self.cbBtnIndexOnPress(None)

    def cbSldDeleteOnRelease(self, widget):
        if widget.value > 0.95:
            self.cbNoFurtherReviewThisWord(None)
            widget.value = 0
        else:
            widget.value = 0

    def cbNextCard(self, iQuality, widget): # iQuality = -1 if there is no need to update record
        global autoPlayThread
        if g_strCurrentOS == "linux":
            autoPlayThread.kill()

        
        self.ChangeWordLearningBtns(0)
        conn = sqlite3.connect(g_strDBPath)
        cursor = conn.cursor()

        global g_bSwitchContinueNewWords, g_iTmpCntStudiedNewWords, g_iTmpCntStudiedOldWords, g_iTmpPrevWordNo, g_iTmpPrevCardType

        today = datetime.today().date()
        strToday = datetime.now().strftime('%Y-%m-%d')

        # Write old word record
        if iQuality != -1:
            if len(self.m_lsRememberSeq) != 0:
                if self.m_lsRememberSeq[0][0]== g_iTmpPrevWordNo and self.m_lsRememberSeq[0][1] == g_iTmpPrevCardType:
                    self.m_lsRememberSeq.pop(0)
            cursor.execute('SELECT * FROM statistics WHERE wordlist=? AND word_no=? AND card_type=? AND review_date IS NULL', \
                           (g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
            result = cursor.fetchall()

            if len(result) != 0:
                #It is a new word
                review = SMTwo.first_review(iQuality, strToday)

                sql = """
                    UPDATE statistics 
                    SET review_date = ?, easiness = ?, interval = ?, repetitions = ? 
                    WHERE wordlist = ? AND word_no = ? AND card_type = ?
                """
                if iQuality == 0:
                    cursor.execute(sql, (today, review.easiness, review.interval, review.repetitions, g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
                else:
                    cursor.execute(sql, (review.review_date, review.easiness, review.interval, review.repetitions, g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
                
            else:
                sql = '''
                    SELECT easiness, interval, repetitions
                    FROM statistics
                    WHERE wordlist = ? AND word_no = ? AND card_type = ?
                '''
                cursor.execute(sql, (g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
                result = cursor.fetchall()
                easiness, interval, repetitions = result[0]   
                review = SMTwo(easiness, interval, repetitions).review(iQuality, strToday)
                sql = """
                    UPDATE statistics 
                    SET review_date = ?, easiness = ?, interval = ?, repetitions = ? 
                    WHERE wordlist = ? AND word_no = ? AND card_type = ?
                """
                if iQuality == 0:
                    cursor.execute(sql, (today, review.easiness, review.interval, review.repetitions, g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
                else:
                    cursor.execute(sql, (review.review_date, review.easiness, review.interval, review.repetitions, g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
        if iQuality == 5:
            # Delete card = Use an huge date
            sql = """
                UPDATE statistics
                SET review_date = '9999-12-31'
                WHERE wordlist = ? AND word_no = ? AND card_type = ?
            """
            cursor.execute(sql,(g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
        # Update tag
        cursor.execute("SELECT COUNT(*) FROM statistics WHERE wordlist = ? AND review_date IS NOT NULL",(g_strOpenedWordListName,))
        iCntStudiedCards = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM statistics WHERE wordlist = ? AND review_date IS NULL",(g_strOpenedWordListName,))
        iCntRestNewCards = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM statistics WHERE wordlist = ? AND review_date <= ?", (g_strOpenedWordListName, strToday))
        iCntDueTimeCards = cursor.fetchone()[0]

        self.lblLearningProgressValue_boxIndexBody.text = str(iCntStudiedCards)+"("+ str(g_iTmpCntStudiedNewWords) + "," + str(g_iTmpCntStudiedOldWords) \
            + ")/" +str(iCntRestNewCards) + "/"+str(iCntDueTimeCards)

        bFindAnOldWord = False
        # Update the learning process text
        if g_bOnlyReview == False:
            if iCntDueTimeCards != 0:
                if iCntRestNewCards != 0:
                    # Get if there is no card needing reviewing
                    if g_dictWordLists[g_strOpenedWordListName]['policy'] == 0:
                        # Learn new words first
                        if g_iTmpCntStudiedNewWords == 0:
                                bFindAnOldWord = False
                                g_bSwitchContinueNewWords = True
                        else:
                            if (g_iTmpCntStudiedNewWords % g_dictWordLists[g_strOpenedWordListName]['NewWordsPerGroup'] == 0)\
                                and g_bSwitchContinueNewWords == True:
                                bFindAnOldWord = False
                                g_bSwitchContinueNewWords = False
                            elif g_iTmpCntStudiedOldWords % g_dictWordLists[g_strOpenedWordListName]['OldWordsPerGroup'] == 0 \
                                and g_bSwitchContinueNewWords == False:
                                bFindAnOldWord = True
                                g_bSwitchContinueNewWords = True
                            else:
                                bFindAnOldWord = not g_bSwitchContinueNewWords
                    elif g_dictWordLists[g_strOpenedWordListName]['policy'] == 1:
                        # Randomly
                        iRandom = random.randint(1, g_dictWordLists[g_strOpenedWordListName]['NewWordsPerGroup'] + g_dictWordLists[g_strOpenedWordListName]['OldWordsPerGroup'])
                        if iRandom < g_dictWordLists[g_strOpenedWordListName]['NewWordsPerGroup']:
                            bFindAnOldWord = True
                        else:
                            bFindAnOldWord = False
                    elif g_dictWordLists[g_strOpenedWordListName]['policy'] == 2:
                        # Review first
                        if g_iTmpCntStudiedOldWords == 0:
                            bFindAnOldWord = True
                            g_bSwitchContinueNewWords = False
                        else:
                            if (g_iTmpCntStudiedOldWords % g_dictWordLists[g_strOpenedWordListName]['OldWordsPerGroup'] == 0)\
                                and g_bSwitchContinueNewWords == False:
                                bFindAnOldWord == True
                                g_bSwitchContinueNewWords = True
                            elif (g_iTmpCntStudiedNewWords % g_dictWordLists[g_strOpenedWordListName]['NewWordsPerGroup'] == 0)\
                                and g_bSwitchContinueNewWords == True:
                                bFindAnOldWord = False
                                g_bSwitchContinueNewWords = False
                            else:
                                bFindAnOldWord = not g_bSwitchContinueNewWords
                else: # There is no new cards
                    bFindAnOldWord = True
            else: # All card has been reviewed
                bFindAnOldWord = False
        else:
            bFindAnOldWord = True
        
        if bFindAnOldWord == True:
            conn.create_function('power', 2, lambda _x, _y: _x**_y) #SQLite does not support exponentiation 
            sql = """
                SELECT word_no, card_type 
                FROM statistics
                WHERE wordlist = ? 
                    AND review_date IS NOT NULL 
                    AND review_date <= ?
                ORDER BY review_date ASC, ( power(easiness, repetitions) / interval) ASC
                LIMIT 1;
            """
            cursor.execute(sql, (g_strOpenedWordListName, strToday))
            result = cursor.fetchone()
            
            if result == None:
                if g_strCurrentOS != 'macOS':
                    self.wbWord_boxIndexBody.set_content("NO_CONTENT", g_EmptyHtml)
                else:
                    self.wbWord_boxIndexBody.url = "http://127.0.0.1"
                
                conn.commit()
                conn.close()
                # If only review is true and rest new cards is not 0,
                if g_bOnlyReview == True and iCntRestNewCards > 0:
                    if g_iTmpCntStudiedNewWords == 0 and g_iTmpCntStudiedOldWords == 0: # A bug: if the program just runs, the info dialog will raise RuntimeError.
                        return
                    self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_ContinueToStudyNewWords)
                    return
                # All old words have been studied.
                if g_iTmpCntStudiedNewWords == 0 and g_iTmpCntStudiedOldWords == 0: # A bug: if the program just runs, the info dialog will raise RuntimeError.
                    return
                self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_MissionComplete)
                return
            else:
                g_iTmpPrevWordNo = result[0]
                g_iTmpPrevCardType = result[1]
                g_iTmpCntStudiedOldWords += 1
            
            sql = '''
                SELECT easiness, interval, repetitions
                FROM statistics
                WHERE wordlist = ? AND word_no = ? AND card_type = ?
            '''
            cursor.execute(sql, (g_strOpenedWordListName, g_iTmpPrevWordNo, g_iTmpPrevCardType))
            result = cursor.fetchall()
            easiness, interval, repetitions = result[0]

            review = SMTwo(easiness, interval, repetitions).review(0, strToday)
            self.btnStrange.text = "😞 (0 d)"

            review2 = SMTwo(easiness, interval, repetitions).review(2, strToday)
            self.btnVague.text = "😐 ("+str((review2.review_date - today).days)+" d)"

            review3 = SMTwo(easiness, interval, repetitions).review(4, strToday)
            self.btnFamiliar.text = "😄 ("+str((review3.review_date - today).days)+" d)"

        else:
            #Find a new word
            if len(self.m_lsRememberSeq) == 0:
                # All new words have been studied. 
                if g_strCurrentOS != 'macOS':
                    self.wbWord_boxIndexBody.set_content("NO_CONTENT", g_EmptyHtml)
                else:
                    self.wbWord_boxIndexBody.url = "http://127.0.0.1"

                conn.commit()
                conn.close()
                if g_iTmpCntStudiedNewWords == 0 and g_iTmpCntStudiedOldWords == 0: # A bug: if the program just runs, the info dialog will raise RuntimeError.
                    return
                self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_MissionComplete)
                return
            else:
                g_iTmpPrevWordNo = self.m_lsRememberSeq[0][0]
                g_iTmpPrevCardType = self.m_lsRememberSeq[0][1]
                g_iTmpCntStudiedNewWords += 1
            
            review = SMTwo.first_review(0, strToday)
            self.btnStrange.text = "😞 (0 d)"
            
            review2 = SMTwo.first_review(2, strToday)
            self.btnVague.text = "😐 ("+str((review2.review_date - today).days)+" d)"
            
            review3 = SMTwo.first_review(4, strToday)
            self.btnFamiliar.text = "😄 ("+str((review3.review_date - today).days)+" d)"
        
        conn.commit()
        conn.close()

        dictCardTypeThisCardUses = g_dictCards[g_dictWordLists[g_strOpenedWordListName]['CardType'+str(g_iTmpPrevCardType)]]
        if len(dictCardTypeThisCardUses) == 1:
            #Display web
            xlsx = load_workbook(g_dictWordLists[g_strOpenedWordListName]['FilePath'])
            sheet = xlsx.worksheets[0]
            dicDynVariables = {}
            for i in range(1, sheet.max_column + 1):
                dicDynVariables[ConvertNumToExcelColTitle(i)] = FilterForExcelNoneValue(sheet.cell(row = g_iTmpPrevWordNo, column = i).value)
            
            if g_strCurrentOS != "macOS":
                strRenderedHTML = GetRenderedHTML(
                    dictCardTypeThisCardUses['single-sided']['frontend'],
                    dictCardTypeThisCardUses['single-sided']['backend'],
                    dicDynVariables
                )
                
                if g_strCurrentOS == "linux":
                    autoPlayThread = AutoPlayThread(strRenderedHTML)
                    autoPlayThread.setDaemon(True)
                    autoPlayThread.start()

                self.wbWord_boxIndexBody.set_content("wordgets", strRenderedHTML)
            else:
                
                flaskThreadOnMac = FlaskThreadOnMac(dictCardTypeThisCardUses['single-sided']['frontend'], dictCardTypeThisCardUses['single-sided']['backend'], dicDynVariables)
                flaskThreadOnMac.setDaemon(True)
                flaskThreadOnMac.start()
                self.wbWord_boxIndexBody.url = "http://127.0.0.1"
                self.wbWord_boxIndexBody.url = "http://127.0.0.1:5000"
                flaskThreadOnMac.kill()

            #Change button
            self.ChangeWordLearningBtns(2)
        else:
            #Display web
            xlsx = load_workbook(g_dictWordLists[g_strOpenedWordListName]['FilePath'])
            sheet = xlsx.worksheets[0]
            dicDynVariables = {}
            for i in range(1, sheet.max_column + 1):
                dicDynVariables[ConvertNumToExcelColTitle(i)] = FilterForExcelNoneValue(sheet.cell(row = g_iTmpPrevWordNo, column = i).value)
            
            if g_strCurrentOS != 'macOS':
                strRenderedHTML = GetRenderedHTML(
                    dictCardTypeThisCardUses['front']['frontend'],
                    dictCardTypeThisCardUses['front']['backend'],
                    dicDynVariables
                )

                if g_strCurrentOS == "linux":
                    autoPlayThread = AutoPlayThread(strRenderedHTML)
                    autoPlayThread.setDaemon(True)
                    autoPlayThread.start()

                self.wbWord_boxIndexBody.set_content("wordgets", strRenderedHTML)
            else:
                
                flaskThreadOnMac = FlaskThreadOnMac(dictCardTypeThisCardUses['front']['frontend'], dictCardTypeThisCardUses['front']['backend'], dicDynVariables)
                flaskThreadOnMac.setDaemon(True)
                flaskThreadOnMac.start()
                self.wbWord_boxIndexBody.url = "http://127.0.0.1"
                self.wbWord_boxIndexBody.url = "http://127.0.0.1:5000"
                flaskThreadOnMac.kill()

            #Change button
            self.ChangeWordLearningBtns(1)

    def cbShowBack(self, widget):
        global autoPlayThread
        if g_strCurrentOS == "linux":
            autoPlayThread.kill()
        self.ChangeWordLearningBtns(0)
        dictCardTypeThisCardUses = g_dictCards[g_dictWordLists[g_strOpenedWordListName]['CardType'+str(g_iTmpPrevCardType)]]

        xlsx = load_workbook(g_dictWordLists[g_strOpenedWordListName]['FilePath'])
        sheet = xlsx.worksheets[0]
        dicDynVariables = {}
        for i in range(1, sheet.max_column + 1):
            dicDynVariables[ConvertNumToExcelColTitle(i)] = FilterForExcelNoneValue(sheet.cell(row = g_iTmpPrevWordNo, column = i).value)
        
        if g_strCurrentOS != 'macOS':
            strRenderedHTML = GetRenderedHTML(
                dictCardTypeThisCardUses['back']['frontend'],
                dictCardTypeThisCardUses['back']['backend'],
                dicDynVariables
            )
            if g_strCurrentOS == "linux":
                autoPlayThread = AutoPlayThread(strRenderedHTML)
                autoPlayThread.setDaemon(True)
                autoPlayThread.start()
            
            self.wbWord_boxIndexBody.set_content("wordgets", strRenderedHTML)

        else:
            flaskThreadOnMac = FlaskThreadOnMac(dictCardTypeThisCardUses['back']['frontend'], dictCardTypeThisCardUses['back']['backend'], dicDynVariables)
            flaskThreadOnMac.setDaemon(True)
            flaskThreadOnMac.start()
            self.wbWord_boxIndexBody.url = "http://127.0.0.1"
            self.wbWord_boxIndexBody.url = "http://127.0.0.1:5000"
            flaskThreadOnMac.kill()
        
        self.ChangeWordLearningBtns(2)

    def cbChangeCurWordList(self, widget):
        global g_iTmpPrevWordNo, g_iTmpPrevCardType, g_iTmpCntStudiedNewWords, g_iTmpCntStudiedOldWords, g_strOpenedWordListName
        if widget.value == g_strOpenedWordListName:
            return
        if widget.value == None:
            g_strOpenedWordListName = ""
        else:
            g_strOpenedWordListName = widget.value
        g_iTmpPrevWordNo = -1
        g_iTmpPrevCardType = -1
        g_iTmpCntStudiedNewWords = 0
        g_iTmpCntStudiedOldWords = 0
        self.lblLearningProgressValue_boxIndexBody.text = ""
        self.GenerateNewWordsSeqOfCurrentWordList()
        self.cbNextCard(-1, None)
        self.SaveConfiguration()

    def cbNoFurtherReviewThisWord(self, widget): # async
        self.cbNextCard(5, None)
        #if await self.main_window.question_dialog(title = g_strQuestionDialogTitle, message = g_strQuestionDialogMsg_NoFurtherReview):
        #    self.cbNextCard(5, None)
        #else:
        #    return
    
    def cbChangeChkReviewOnly(self, widget):
        global g_bOnlyReview
        g_bOnlyReview = widget.value
        if len(self.cboCurWordList_boxIndexBody.items) == 0:
            return
        if g_bOnlyReview == False:
            self.cbNextCard(-1, None)
    
    def cbClearCache(self, widget):
        shutil.rmtree(os.path.join(g_strDataPath,"Cache"), True)
        self.lblCacheSize_Row3Col2.text = str(round(GetFolderSize(os.path.join(g_strDataPath, "Cache")) / 1024 / 1024, 2)) + " MB"

    def cbMigrateFromAnkiOnPress(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(self.boxMigration)
        self.txtApkgFilePath_boxMigration.value = ""
        self.cboWordLists_boxMigration.items = list(g_dictWordLists.keys())
        self.prgRateOfMigration_boxMigration.value = 0
        self.lblMigrationInfo_boxMigration.text = ""
    
    async def cbMigrate(self, widget):
        strApkgFilePath = self.txtApkgFilePath_boxMigration.value
        wordlist = self.cboWordLists_boxMigration.value
        if strApkgFilePath == "" and (wordlist == None or wordlist == ""):
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_CanNotMigrate)
            return

        bDeleteApkg = False
        if strApkgFilePath[:4]=="http":
            try:
                response = requests.get(strApkgFilePath, stream = True)
                filename = "anki_dl.apkg"
                strFileWriteTo = os.path.join(g_strDataPath, filename)
                with open(strFileWriteTo, "wb") as f:
                    for chunk in response.iter_content(chunk_size = 1024*1024):
                            if chunk:
                                f.write(chunk)
                strApkgFilePath = strFileWriteTo
                bDeleteApkg = True
            except:
                self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_DownloadFailed)
                return
            
        await self.main_window.info_dialog(g_strInfoDialogTitle_Attention, g_strInfoDialogMsg_DoNotClose)
        
        strApkgFolder = os.path.dirname(strApkgFilePath)
        strApkgFileNameWithExtension = os.path.basename(strApkgFilePath)
        strApkgFileName = os.path.splitext(strApkgFileNameWithExtension)[0]
        strExtractToFolder = os.path.join(
            strApkgFolder,
            strApkgFileName
        )
        os.makedirs(strExtractToFolder, exist_ok=True)

        try:
            self.lblMigrationInfo_boxMigration.text = g_strLblMigrationInfo_Dyn_Step1
            zip_file = zipfile.ZipFile(strApkgFilePath)
            zip_file.extractall(strExtractToFolder)
            strAnki21FilePath = os.path.join(strExtractToFolder,"collection.anki21")
            if not os.path.exists(strAnki21FilePath):
                raise Exception
            self.lblMigrationInfo_boxMigration.text = g_strLblMigrationInfo_Dyn_Step2
            connAnkiDB = sqlite3.connect(strAnki21FilePath)
            cursorAnkiDB = connAnkiDB.cursor()
            cursorAnkiDB.execute('SELECT id, cid, ease, ivl, type FROM revlog')
            lsRevlog = cursorAnkiDB.fetchall()
            cursorAnkiDB.execute('SELECT id, nid, ord, reps FROM cards')
            lsCards = cursorAnkiDB.fetchall()
            cursorAnkiDB.execute('SELECT id, sfld FROM notes')
            lsNotes = cursorAnkiDB.fetchall()
            connAnkiDB.close()
            dictNotes = {item[0]: item[1] for item in lsNotes}
            iLenLsRevlog = len(lsRevlog)
            iCnt = 0
            while iCnt < iLenLsRevlog:
                for i in range(0, iCnt):
                    if lsRevlog[i][1] == lsRevlog[iCnt][1]:
                        lsRevlog = lsRevlog[:i]+lsRevlog[i+1:]
                        iCnt-=1
                        iLenLsRevlog = len(lsRevlog)
                        break
                iCnt+=1
            dictMergedCardsAndNotes = {}
            for i in range(len(lsCards)):
                dictMergedCardsAndNotes[lsCards[i][0]] = [dictNotes[lsCards[i][1]], lsCards[i][2], lsCards[i][3]]
            
            self.lblMigrationInfo_boxMigration.text = g_strLblMigrationInfo_Dyn_Step3
            strWordListFilePath = g_dictWordLists[wordlist]['FilePath']
            xlsx = load_workbook(strWordListFilePath, read_only = True)
            sheet = xlsx.worksheets[0]
            iRow = 2
            dictWordID = {}
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                self.prgRateOfMigration_boxMigration.value = (i - 1)/(sheet.max_row- 1)
                dictWordID[row[0]] = iRow
                iRow+=1
            self.prgRateOfMigration_boxMigration.value = 0

            self.lblMigrationInfo_boxMigration.text = g_strLblMigrationInfo_Dyn_Step4
            lsMigrationContents = []
            for i in range(len(lsRevlog)):
                self.prgRateOfMigration_boxMigration.value = (i+1)/len(lsRevlog)
                id = lsRevlog[i][0]
                cid = lsRevlog[i][1]
                cid_values = dictMergedCardsAndNotes[cid]
                sfld = cid_values[0]
                ord = cid_values[1]
                reps = cid_values[2]
                ease = lsRevlog[i][2]
                ivl = lsRevlog[i][3]
                type_ = lsRevlog[i][4]
                word = sfld
                word_no = -1
                try:
                    word_no = dictWordID[word]
                except:
                    pass
                card_type = ord + 1
                review_date_num = id
                interval = -1
                if ivl >= 0:
                    review_date_num += ivl * 24 * 60 * 60 * 1000
                    interval = ivl
                else:
                    review_date_num += -ivl * 1000
                    interval = 0
                timestamp = review_date_num / 1000
                dt = datetime.fromtimestamp(timestamp)
                review_date = dt.strftime('%Y-%m-%d')
                repetitions = reps
                easiness = -1
                if type_ == 0 or type_ == 2:
                    if ease == 1:
                        easiness = 0
                    elif ease == 2:
                        easiness = 2
                    elif ease >= 3:
                        easiness =4
                else:
                    if ease == 1:
                        easiness = 0
                    elif ease == 2 or ease == 3:
                        easiness = 2
                    elif ease >= 4:
                        easiness = 4
                if word_no == -1:
                    continue
                lsMigrationContents.append(
                    (
                    word_no,
                    card_type,
                    review_date,
                    easiness,
                    interval,
                    repetitions)
                )
            self.prgRateOfMigration_boxMigration.value = 0

            self.lblMigrationInfo_boxMigration.text = g_strLblMigrationInfo_Dyn_Step5

            connWordgetDB = sqlite3.connect(g_strDBPath)
            cursorWordgetDB = connWordgetDB.cursor()

            for i in range(len(lsMigrationContents)):
                self.prgRateOfMigration_boxMigration.value = (i+1)/len(lsMigrationContents)

                word_no = lsMigrationContents[i][0]
                card_type = lsMigrationContents[i][1]
                review_date = lsMigrationContents[i][2]
                easiness = lsMigrationContents[i][3]
                interval = lsMigrationContents[i][4]
                repetitions = lsMigrationContents[i][5]

                cursorWordgetDB.execute("UPDATE statistics SET review_date=?, easiness=?, interval=?, repetitions=? WHERE wordlist=? AND word_no=? AND card_type=?", 
                    (review_date, easiness, interval, repetitions, wordlist, word_no, card_type))
            connWordgetDB.commit()
            connWordgetDB.close()
            await self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_MigrationCompleted)

        except:
            await self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_MigrationFailed)

        self.lblMigrationInfo_boxMigration.text = g_strLblMigrationInfo_Dyn_Step6
        try:
            if os.path.exists(strExtractToFolder) == True and strExtractToFolder != "":
                shutil.rmtree(strExtractToFolder)
            
        except:
            pass
            
        self.lblMigrationInfo_boxMigration.text = ""
        self.prgRateOfMigration_boxMigration.value = 0
        
    # Online services---------------------------------------------------------------------------------------------------------------------------------------
    def cbLoginCloudSvc(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.boxBody.add(
            self.boxCloudSvcP1
        )
    
    def cbSetCloudSvc(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.txtCloudSvcP2AppId_boxCloudSvcP2.value = g_strAppId
        self.boxBody.add(
            self.boxCloudSvcP2
        )
    ''' 
    def NewThreadToDetectJump(self):   #unstable 
        while True:
            if self.boxCloudSvcP3 not in self.boxBody.children:
                return
            if self.wbCloudSvcP3_LoginAccount.url.find("openapi.baidu.com/oauth/2.0/login_success") >= 0:
                url = self.wbCloudSvcP3_LoginAccount.url
                query_args = urllib.parse.parse_qs(url.split("#")[1])
                access_token = query_args['access_token'][0]
                #self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_LoginSuccess)
                #Save AppID and access_token
                global g_strAppId, g_strAccessToken, g_strUsername
                g_strAppId = self.m_strTmpAppId
                g_strAccessToken = access_token
                url = f"https://pan.baidu.com/rest/2.0/xpan/nas?access_token={access_token}&method=uinfo"
                payload = {}
                response = requests.request("GET", url, data = payload)
                strResponse = response.text.encode('utf8').decode()
                dictResponse = eval(strResponse)
                username = dictResponse['baidu_name']
                g_strUsername = username

                self.lblLoginState_Row2Col2.text                = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGIN"] % username)  #Crash when using gloabal variable

                self.btnSync_boxIndexBody.enabled = True

                while len(self.boxBody.children) != 0:
                    self.boxBody.remove(self.boxBody.children[0])
                
                asyncio.run(self.SyncDBToLocal())
                self.SaveConfiguration()
                self.GenerateNewWordsSeqOfCurrentWordList() # It is necessary.
                #self.cbBtnSettingsOnPress(None)
                return
            time.sleep(0.5)
        '''
    
    async def cbLoginOnWebviewLoad(self, widget):  
        if widget.url.find("openapi.baidu.com/oauth/2.0/login_success") >= 0:
            url = widget.url
            query_args = urllib.parse.parse_qs(url.split("#")[1])
            access_token = query_args['access_token'][0]
            #self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_LoginSuccess)
            #Save AppID and access_token
            global g_strAppId, g_strAccessToken, g_strUsername
            g_strAppId = self.m_strTmpAppId
            g_strAccessToken = access_token
            url = f"https://pan.baidu.com/rest/2.0/xpan/nas?access_token={access_token}&method=uinfo"
            payload = {}
            response = requests.request("GET", url, data = payload)
            strResponse = response.text.encode('utf8').decode()
            dictResponse = eval(strResponse)
            username = dictResponse['baidu_name']
            g_strUsername = username

            self.lblLoginState_Row2Col2.text                = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGIN"] % username)  #Crash when using gloabal variable

            self.btnSync_boxIndexBody.enabled = True

            #while len(self.boxBody.children) != 0:
            #    self.boxBody.remove(self.boxBody.children[0])

            await self.SyncDBToLocal(None)
            self.SaveConfiguration()
            self.GenerateNewWordsSeqOfCurrentWordList() # It is necessary.
            self.cbBtnSettingsOnPress(None)
            
    async def cbManualVerificationOnPress(self, widget):
        if self.wbCloudSvcP3_LoginAccount.url.find("openapi.baidu.com/oauth/2.0/login_success") >= 0:
            url = self.wbCloudSvcP3_LoginAccount.url
            query_args = urllib.parse.parse_qs(url.split("#")[1])
            access_token = query_args['access_token'][0]
            global g_strAppId, g_strAccessToken, g_strUsername
            g_strAppId = self.m_strTmpAppId
            g_strAccessToken = access_token
            url = f"https://pan.baidu.com/rest/2.0/xpan/nas?access_token={access_token}&method=uinfo"
            payload = {}
            response = requests.request("GET", url, data = payload)
            strResponse = response.text.encode('utf8').decode()
            dictResponse = eval(strResponse)
            username = dictResponse['baidu_name']
            g_strUsername = username

            self.lblLoginState_Row2Col2.text                = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGIN"] % username)  #Crash when using gloabal variable
            self.btnSync_boxIndexBody.enabled = True
            
            await self.SyncDBToLocal(None)
            self.SaveConfiguration()
            self.GenerateNewWordsSeqOfCurrentWordList() # It is necessary.
            self.cbBtnSettingsOnPress(None)
        else:
            self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_VerificationFailed)

    def cbLoginAccount(self, widget):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.m_strTmpAppId = self.txtCloudSvcP2AppId_boxCloudSvcP2.value
        self.txtCloudSvcP2AppId_boxCloudSvcP2.value = ""
        self.boxBody.add(
            self.boxCloudSvcP3
        )
        url = f"https://openapi.baidu.com/oauth/2.0/authorize?response_type=token&client_id={self.m_strTmpAppId}&redirect_uri=oob&scope=basic,netdisk&display=mobile"
        
        if g_strCurrentOS == "windows" or g_strCurrentOS == "macOS": #Linux: crashed; Android: not implemented.==========================================================
            self.wbCloudSvcP3_LoginAccount.on_webview_load = self.cbLoginOnWebviewLoad
        self.wbCloudSvcP3_LoginAccount.url = url
        
        #thread = Thread(target=self.NewThreadToDetectJump)
        #thread.start()
        #while True:
        #    if g_bTmpLoginSync == True:
        #        asyncio.run(self.SyncDBToLocal())
        #        break
        #    time.sleep(0.5)
    
    async def cbLogoutAccount(self, widget):
        global g_strAccessToken, g_strUsername
        if await self.main_window.question_dialog(title = g_strQuestionDialogTitle, message = g_strQuestionDialogMsg_Logout):
            g_strAccessToken = ""
            g_strUsername = ""

            self.lblLoginState_Row2Col2.text                = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGOUT"]
            self.btnSync_boxIndexBody.enabled = False
            self.cbBtnSettingsOnPress(None)
        else:
            return

    def ChangeAccountBtnState(self):
        if g_strUsername != "":
            self.btnLogin_Row2Col2.text                         = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_ANOTHERLOGIN_TEXT']
            self.btnLogout_Row2Col2.style.update(visibility = VISIBLE)
        else:
            self.btnLogin_Row2Col2.text                         = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_LOGIN_TEXT']
            self.btnLogout_Row2Col2.style.update(visibility = HIDDEN)

    def GetCloudDBFileID(self):
        url = f"https://pan.baidu.com/rest/2.0/xpan/file?method=list&dir=/wordgets_sync/&order=time&start=0&limit=100&web=web&folder=0&access_token={g_strAccessToken}&desc=1"
        payload = {}
        files = {}
        headers = {
            'User-Agent': 'pan.baidu.com'
        }
        response = requests.request("GET", url, headers=headers, data = payload, files = files)
        strResponse = response.text.encode('utf8').decode().replace("\\/","/")
        dictResponse = eval(strResponse)
        errno = dictResponse['errno']
        if errno == 0: #Folder exists
            for eachFile in dictResponse['list']:
                if eachFile['server_filename'] == 'wordgets_stat.db':
                    return eachFile['fs_id']   # FileID
            return 0
        elif errno == -9: #Folder does not exist
            url = f"https://pan.baidu.com/rest/2.0/xpan/file?method=create&access_token={g_strAccessToken}"
            payload = {
                'path': '/wordgets_sync',
                'rtype': '0',
                'isdir': '1'
            }
            files = [
            ]
            headers = {
            }
            response = requests.request("POST", url, headers=headers, data = payload, files = files)
            return 0
        else: #Include: Access_token overdue; over-frequent request
            return -1

    def DownloadDBFromSyncAndMergeRecord(self, iFileID):
        #Query file information
        url = f"http://pan.baidu.com/rest/2.0/xpan/multimedia?method=filemetas&access_token={g_strAccessToken}&fsids=[{iFileID}]&thumb=1&dlink=1&extra=1"

        payload = {}
        files = {}
        headers = {
            'User-Agent': 'pan.baidu.com'
        }

        response = requests.request("GET", url, headers=headers, data = payload, files = files)

        strResponse = response.text.encode('utf8').decode().replace("\\/","/")

        dictResponse = eval(strResponse)
        errno = dictResponse['errno']
        
        if errno != 0:
            return -1
        dlink = dictResponse['list'][0]['dlink']
        #md5 = dictResponse['list'][0]['md5']

        #Download file
        url = dlink + f"&access_token={g_strAccessToken}"
        payload = {}

        files = {}
        headers = {
            'User-Agent': 'pan.baidu.com'
        }

        response = requests.request("GET", url, headers=headers, data = payload, files = files)
        file_binary = response.content   # not the so-called .text on the website
        with open(g_strDBPath+"_tmp", "wb") as f: 
            f.write(file_binary)

        # Baidu Netdisk uses different md5 policy. Directly try to load database
        try:
            if os.path.exists(g_strDBPath) == False:
                os.rename(g_strDBPath+"_tmp", g_strDBPath)
                return 1
            conn1 = sqlite3.connect(g_strDBPath)
            cursor1 = conn1.cursor()
            cursor1.execute("SELECT * FROM statistics ORDER BY wordlist, word_no, card_type")
            data1 = cursor1.fetchall()
            conn1.close()

            conn2 = sqlite3.connect(g_strDBPath+"_tmp")
            cursor2 = conn2.cursor()
            cursor2.execute("SELECT * FROM statistics ORDER BY wordlist, word_no, card_type")
            data2 = cursor2.fetchall()
            conn2.close()

            if len(data1) != len(data2):
                return 0 
            bOneToOneCorrespondence = True
            for row1, row2 in zip(data1, data2):
                if row1[:3] != row2[:3]:
                    bOneToOneCorrespondence = False
                    break
            if bOneToOneCorrespondence == False:
                return 0
            
            # Update according date
            conn1 = sqlite3.connect(g_strDBPath)
            cursor1 = conn1.cursor()
            for i in range(len(data2)):
                wordlist = data2[i][0]
                word_no = data2[i][1]
                card_type = data2[i][2]
                review_date = data2[i][3]
                easiness = data2[i][4]
                interval = data2[i][5]
                repetitions = data2[i][6]
                if review_date == None or review_date == "":
                    continue

                sql_query = '''
                UPDATE statistics
                SET review_date = ?, easiness = ?, interval = ?, repetitions = ?
                WHERE (wordlist = ? AND word_no = ? AND card_type = ?)
                AND (review_date <= ? OR review_date IS NULL)
                '''
                data = (review_date, easiness, interval, repetitions, wordlist, word_no, card_type, review_date)
                conn1.execute(sql_query, data)
            conn1.commit()
            conn1.close()

            # Delete temporary file
            os.remove(g_strDBPath+"_tmp")

            return 1
        except Exception:
            return -2   #Broken downloaded file 

    def UploadDBFile(self):
        global g_iTmpSyncState
        try:
            #In case of non-existed folder
            url = f"https://pan.baidu.com/rest/2.0/xpan/file?method=create&access_token={g_strAccessToken}"
            payload = {
                    'path': '/wordgets_sync',
                    'rtype': '0',
                    'isdir': '1'
                }
            files = [
                ]
            headers = {
                }
            response = requests.request("POST", url, headers=headers, data = payload, files = files)

            #Preupload
            url = f"http://pan.baidu.com/rest/2.0/xpan/file?method=precreate&access_token={g_strAccessToken}"
            shutil.copy2(g_strDBPath, g_strDBPath + "_tmp")
            lsChunks = []
            lsMd5s =[]
            with open(g_strDBPath + "_tmp", 'rb') as file:
                chunk_index = 0
                while True:
                    chunk = file.read(4 * 1024 * 1024)    #4M/chunk
                    if not chunk:
                        break
                    chunk_md5 = hashlib.md5(chunk).hexdigest()
                    lsChunks.append(chunk)
                    lsMd5s.append(chunk_md5)
                    chunk_index += 1
            
            iDBFileSize = os.path.getsize(g_strDBPath + "_tmp")

            payload = {'path': '/wordgets_sync/wordgets_stat.db_tmp',
                'size': str(iDBFileSize),
                'rtype': '3',
                'isdir': '0',
                'autoinit': '1',
                'block_list': str(lsMd5s).replace("'","\"")
            }
            files = [
            ]
            headers = {
            }

            response = requests.request("POST", url, headers=headers, data = payload, files = files)
            strResponse = response.text.encode('utf8').decode().replace("\\/","/")
            
            dictResponse = eval(strResponse)
        
            uploadid = dictResponse['uploadid']
            # Upload database
            for i in range(len(lsChunks)):
                url = f"https://d.pcs.baidu.com/rest/2.0/pcs/superfile2?method=upload&access_token={g_strAccessToken}&path=/wordgets_sync/wordgets_stat.db_tmp&type=tmpfile&uploadid={uploadid}&partseq=" + str(i)
                
                payload = {}
                files = [
                    ('file',lsChunks[i])
                ]
                headers = {
                }

                response = requests.request("POST", url, headers=headers, data = payload, files = files)
            
            # Create / Merge
            url = f"https://pan.baidu.com/rest/2.0/xpan/file?method=create&access_token={g_strAccessToken}"

            payload = {'path': '/wordgets_sync/wordgets_stat.db_tmp',
                'size': str(iDBFileSize),
                'rtype': '3',
                'isdir': '0',
                'uploadid': uploadid,
                'block_list':  str(lsMd5s).replace("'","\"")
            }
            files = [

            ]
            headers = {
            }

            response = requests.request("POST", url, headers=headers, data = payload, files = files)
            # Overwrite cloud database
            
            url = f"https://pan.baidu.com/rest/2.0/xpan/file?method=filemanager&access_token={g_strAccessToken}&opera=rename"

            payload = {
                'async': '2',
                'filelist': '[{"path":"/wordgets_sync/wordgets_stat.db_tmp","newname":"wordgets_stat.db","ondup":"overwrite"}]'
            }

            response = requests.request("POST", url, data = payload)

            g_iTmpSyncState = 1

            self.app.add_background_task(
                self.ChangeSyncStateAgent
            )
        except Exception:
            g_iTmpSyncState = 2
            self.app.add_background_task(
                self.ChangeSyncStateAgent
            )
        finally:
            time.sleep(10)
            g_iTmpSyncState = 0
            self.app.add_background_task(
                self.ChangeSyncStateAgent
            )
            
    def ChangeSyncStateAgent(self, widget):
        if g_iTmpSyncState == -1:
            self.ChangeSyncBtnText(-1)
            self.btnSync_boxIndexBody.enabled = False
            self.btnIndex_boxNavigateBar.enabled = False
            self.btnSettings_boxNavigateBar.enabled = False
            self.btnLibrary_boxNavigateBar.enabled = False
        elif g_iTmpSyncState == 0:
            self.ChangeSyncBtnText(0)
            self.btnSync_boxIndexBody.enabled = True
            self.btnIndex_boxNavigateBar.enabled = True
            self.btnSettings_boxNavigateBar.enabled = True
            self.btnLibrary_boxNavigateBar.enabled = True
        elif g_iTmpSyncState == 1:
            self.ChangeSyncBtnText(1)
        elif g_iTmpSyncState == 2:
            self.ChangeSyncBtnText(2)


    async def SyncDBToLocal(self, widget): #Only work on account login or launching
        while True:
            try:
                global g_strAccessToken, g_strUsername
                iFileId = self.GetCloudDBFileID()
                if iFileId == 0: # New account
                    self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_NotSyncYet)
                    return
                elif iFileId == -1: # Login failed.
                    self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_LoginFailure)
                    g_strAccessToken = ""
                    g_strUsername = ""
                    self.lblLoginState_Row2Col2.text                = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGOUT"]

                    self.btnSync_boxIndexBody.enabled = False

                    self.cbBtnSettingsOnPress(None)
                    return
                iDBChangeState = self.DownloadDBFromSyncAndMergeRecord(iFileId)
                if iDBChangeState == -1:  #Cannot get the download link
                    if await self.main_window.question_dialog(g_strQuestionDialogTitle, g_strQuestionDialogMsg_NullDLink) == True:
                        continue
                    else:
                        return
                elif iDBChangeState == 0:  #Incompatible DB
                    if g_strCurrentOS == 'macOS' and g_bTmpProgramInitEnd == False:
                        self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_LoginFailure)
                        g_strAccessToken = ""
                        g_strUsername = ""
                        self.lblLoginState_Row2Col2.text                = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGOUT"]
                        self.btnSync_boxIndexBody.enabled = False
                        self.cbBtnSettingsOnPress(None)
                        return
                    else:
                        self.main_window.error_dialog(g_strErrorDialogTitle, g_strErrorDialogMsg_IncompatibleDB)
                    self.cbBtnLibraryOnPress(None) 
                    return
                elif iDBChangeState == 1: #Success
                    
                    self.main_window.info_dialog(g_strInfoDialogTitle, g_strInfoDialogMsg_SyncCompleted)
                    return
                else: #-2 File broken
                    if await self.main_window.question_dialog(g_strQuestionDialogTitle, g_strQuestionDialogMsg_BrokenCloudDB) == True:
                        continue
                    else:
                        return
            except Exception: #Network Error
                if await self.main_window.question_dialog(g_strQuestionDialogTitle, g_strQuestionDialogMsg_Disconnected) == True:
                    continue
                else:
                    return

    def cbSyncDBToCloud(self, widget): # Only work on sync button on press
        global g_iTmpSyncState
        g_iTmpSyncState = -1
        self.ChangeSyncStateAgent(None)
        
        thread = Thread(target = self.UploadDBFile)
        thread.start()

    # Online services End----------------------------------------------------------------------------------------------------------------------

    # Functions
    def ChangeLangAccordingToDefaultLang(self):
        self.btnIndex_boxNavigateBar.text                   = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_INDEX_TEXT']
        self.btnSettings_boxNavigateBar.text                = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_SETTINGS_TEXT']
        self.btnLibrary_boxNavigateBar.text                 = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_LIBRARY_TEXT']
        self.lblLang_boxSettingsBody.text                   = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_LANG_TEXT']
        self.lblCards_boxLibraryBody.text                   = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CARDS_TEXT']
        self.lblWordLists_boxLibraryBody.text               = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_WORDLISTS_TEXT']
        self.btnApplyChanges_boxLibraryBody.text            = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_APPLYCHANGES_TEXT']
        self.lblEditCard_boxEditCardBody.text               = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_EDITCARD_TEXT']
        self.lblCardName_boxEditCardBody.text               = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CARDNAME_TEXT']
        self.txtCardName_boxEditCardBody.placeholder        = self.m_dictLanguages[g_strDefaultLang]['STR_TXT_PLACEHOLDER_LEGALNAME']
        self.lblCardFrontFrontend_boxEditCardBody.text      = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CARDFRONT_FRONTEND_TEXT']
        self.lblCardFrontBackend_boxEditCardBody.text       = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CARDFRONT_BACKEND_TEXT']
        self.lblCardBackFrontend_boxEditCardBody.text       = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CARDBACK_FRONTEND_TEXT']
        self.lblCardBackBackend_boxEditCardBody.text        = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CARDBACK_BACKEND_TEXT']
        self.chkEnableCardBack.text                         = self.m_dictLanguages[g_strDefaultLang]['STR_CHK_ENABLECARDBACK_TEXT']
        self.btnSaveCard_boxEditCardBody.text               = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_SAVECARD_TEXT']
        self.lblEditWordlist_boxEditWordListBody.text       = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_EDITWORDLIST_TEXT']
        self.lblWordListName_boxEditWordListBody.text       = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_WORDLISTNAME_TEXT']
        self.txtWordListName_boxEditWordListBody.placeholder= self.m_dictLanguages[g_strDefaultLang]['STR_TXT_PLACEHOLDER_LEGALNAME']
        self.lblWordListXlsx_boxEditWordListBody.text       = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_WORDLISTXLSX_TEXT']
        self.btnSaveWordList_boxEditWordListBody.text       = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_SAVEWORDLIST_TEXT']
        self.lblLearningProgressItem_boxIndexBody.text      = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_LEARNINGPROGRESSITEM_TEXT']
        self.chkReviewOnly_boxIndexBody.text                = self.m_dictLanguages[g_strDefaultLang]['STR_CHK_REVIEWONLY_TEXT']
        self.btnShowBack.text                               = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_SHOWBACK_TEXT']
        self.lblDeleteTip_boxIndexBody.text                 = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_DELETETIP_TEXT']
        self.lblClearCache_boxSettingsBody.text             = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CLEARCACHE_TEXT']
        self.btnClearCache_Row3Col2.text                    = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_CLEARCACHE_TEXT']

        global g_strFront, g_strBack, g_strOnesided, g_strFrontend, g_strBackend, g_strFilePath, g_strCard1, g_strCard2
        g_strFront                                          = self.m_dictLanguages[g_strDefaultLang]['STR_FRONT']
        g_strBack                                           = self.m_dictLanguages[g_strDefaultLang]['STR_BACK']
        g_strOnesided                                       = self.m_dictLanguages[g_strDefaultLang]['STR_ONESIDED']
        g_strFrontend                                       = self.m_dictLanguages[g_strDefaultLang]['STR_FRONTEND']
        g_strBackend                                        = self.m_dictLanguages[g_strDefaultLang]['STR_BACKEND']
        g_strFilePath                                       = self.m_dictLanguages[g_strDefaultLang]['STR_FILEPATH']
        g_strCard1                                          = self.m_dictLanguages[g_strDefaultLang]['STR_CARD_1']
        g_strCard2                                          = self.m_dictLanguages[g_strDefaultLang]['STR_CARD_2']
        global g_strErrorDialogTitle, g_strErrorDialogMsg_EmptyName, g_strErrorDialogMsg_InvalidFileName, g_strErrorDialogMsg_RepetitiveName,\
            g_strErrorDialogMsg_AssociatedWordList, g_strQuestionDialogTitle, g_strQuestionDialogMsg_Delete, g_strErrorDialogMsg_InvalidFile,\
            g_strErrorDialogMsg_NoWordList, g_strErrorDialogMsg_Card1IsNull, g_strErrorDialogMsg_Card1SameAsCard2,\
            g_strQuestionDialogMsg_NotFoundWordList, g_strInfoDialogTitle, g_strInfoDialogMsg_Complete, g_strErrorDialogMsg_DownloadFailed,\
            g_strInfoDialogMsg_MissionComplete, g_strInfoDialogMsg_ContinueToStudyNewWords, g_strQuestionDialogMsg_NoFurtherReview
        g_strErrorDialogTitle                               = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_TITLE']
        g_strErrorDialogMsg_EmptyName                       = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_EMPTYNAME']
        g_strErrorDialogMsg_RepetitiveName                  = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_REPETITIVENAME']
        g_strErrorDialogMsg_InvalidFileName                 = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_INVALIDFILENAME']
        g_strErrorDialogMsg_AssociatedWordList              = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_ASSOCIATEDWORDLIST']
        g_strQuestionDialogTitle                            = self.m_dictLanguages[g_strDefaultLang]['STR_QUESTIONDIALOG_TITLE']
        g_strQuestionDialogMsg_Delete                       = self.m_dictLanguages[g_strDefaultLang]['STR_QUESTIONDIALOG_MSG_DELETE']
        g_strErrorDialogMsg_InvalidFile                     = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_INVALIDFILE']
        g_strErrorDialogMsg_NoWordList                      = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_NOWORDLIST']
        g_strErrorDialogMsg_Card1IsNull                     = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_CARD1ISNULL']
        g_strErrorDialogMsg_Card1SameAsCard2                = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_CARD1SAMEASCARD2']
        g_strQuestionDialogMsg_NotFoundWordList             = self.m_dictLanguages[g_strDefaultLang]['STR_QUESTIONDIALOG_MSG_NOTFOUNDWORDLIST']
        g_strInfoDialogTitle                                = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_TITLE']
        g_strInfoDialogMsg_Complete                         = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_MSG_COMPLETE']
        g_strErrorDialogMsg_DownloadFailed                  = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_DOWNLOADFAILED']
        g_strInfoDialogMsg_MissionComplete                  = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_MSG_MISSIONCOMPLETE']
        g_strInfoDialogMsg_ContinueToStudyNewWords          = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_MSG_CONTINUETOSTUDYINGNEWWORDS']
        g_strQuestionDialogMsg_NoFurtherReview              = self.m_dictLanguages[g_strDefaultLang]['STR_QUESTIONDIALOG_MSG_NOFURTHERREVIEW']
        global g_strLblNewWordsPerGroup, g_strLblOldWordsPerGroup, g_strLblReviewPolicy, g_strCboReviewPolicy_item_LearnFirst, \
            g_strCboReviewPolicy_item_Random, g_strCboReviewPolicy_item_ReviewFirst
        g_strLblNewWordsPerGroup                            = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_NEWWORDSPERGROUP']
        g_strLblOldWordsPerGroup                            = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_OLDWORDSPERGROUP']
        g_strLblReviewPolicy                                = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_REVIEWPOLICY']
        g_strCboReviewPolicy_item_LearnFirst                = self.m_dictLanguages[g_strDefaultLang]['STR_CBO_REVIEWPOLICY_ITEM_LEARNFIRST']
        g_strCboReviewPolicy_item_Random                    = self.m_dictLanguages[g_strDefaultLang]['STR_CBO_REVIEWPOLICY_ITEM_RANDOM']
        g_strCboReviewPolicy_item_ReviewFirst               = self.m_dictLanguages[g_strDefaultLang]['STR_CBO_REVIEWPOLICY_ITEM_REVIEWFIRST']
        
        # Cloud service
        self.lblCloudSvc_boxSettingsBody.text               = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CLOUDSVC_TEXT']
        #self.btnLogin_Row2Col2.text                         = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_LOGIN_TEXT']
        self.lblCloudSvcP1Title_boxCloudSvcP1.text          = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CLOUDSVCP1TITLE_TEXT']
        self.btnCloudSvcP1_BaiduNetdisk_boxCloudSvcP1.text  = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_CLOUDSVCP1_BAIDUNETDISK_TEXT']
        self.lblCloudSvcP2Title_boxCloudSvcP2.text          = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CLOUDSVCP2TITLE_TEXT']
        self.btnCloudSvcP2WhatIsThis_boxCloudSvcP2.text     = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_CLOUDSVCP2WHATISTHIS_TEXT']
        self.btnCloudSvcP2Next_boxCloudSvcP2.text           = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_CLOUDSVCP2NEXT_TEXT']
        self.btnLogout_Row2Col2.text                        = self.m_dictLanguages[g_strDefaultLang]['STR_BTN_LOGOUT_TEXT']
        self.lblCloudSvcP3Title_boxCloudSvcP3.text          = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_CLOUDSVCP3TITLE_TEXT']
        global g_strInfoDialogMsg_LoginSuccess, g_strQuestionDialogMsg_Logout, g_strInfoDialogMsg_LoginFailure, g_strQuestionDialogMsg_Disconnected,\
            g_strInfoDialogMsg_NotSyncYet, g_strErrorDialogMsg_IncompatibleDB, g_strQuestionDialogMsg_NullDLink, g_strQuestionDialogMsg_BrokenCloudDB,\
            g_strInfoDialogMsg_AboutAppId, g_strInfoDialogMsg_SyncCompleted, g_strErrorDialogMsg_VerificationFailed
        g_strInfoDialogMsg_LoginSuccess                     = self.m_dictLanguages[g_strDefaultLang]["STR_INFODIALOG_MSG_LOGINSUCCESS"]
        g_strQuestionDialogMsg_Logout                       = self.m_dictLanguages[g_strDefaultLang]["STR_QUESTIONDIALOG_MSG_LOGOUT"]
        if g_strUsername != "":
            self.lblLoginState_Row2Col2.text                = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGIN"] % g_strUsername
        else:
            self.lblLoginState_Row2Col2.text                = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_LOGINSTATE_TEXT_LOGOUT"]
        self.ChangeAccountBtnState()
        g_strInfoDialogMsg_LoginFailure                     = self.m_dictLanguages[g_strDefaultLang]["STR_INFODIALOG_MSG_LOGINFAILURE"]
        g_strQuestionDialogMsg_Disconnected                 = self.m_dictLanguages[g_strDefaultLang]["STR_QUESTIONDIALOG_MSG_DISCONNECTED"]
        g_strInfoDialogMsg_NotSyncYet                       = self.m_dictLanguages[g_strDefaultLang]["STR_INFODIALOG_MSG_NOTSYNCYET"]
        g_strErrorDialogMsg_IncompatibleDB                  = self.m_dictLanguages[g_strDefaultLang]["STR_ERRORDIALOG_MSG_INCOMPATIBLEDB"]
        g_strQuestionDialogMsg_NullDLink                    = self.m_dictLanguages[g_strDefaultLang]["STR_QUESTIONDIALOG_MSG_NULLDLINK"]
        g_strQuestionDialogMsg_BrokenCloudDB                = self.m_dictLanguages[g_strDefaultLang]["STR_QUESTIONDIALOG_MSG_BROKENCLOUDDB"]
        g_strInfoDialogMsg_AboutAppId                       = self.m_dictLanguages[g_strDefaultLang]["STR_INFODIALOG_MSG_ABOUTAPPID"]
        self.ChangeSyncBtnText(0)
        g_strInfoDialogMsg_SyncCompleted                    = self.m_dictLanguages[g_strDefaultLang]["STR_INFODIALOG_MSG_SYNCCOMPLETED"]
        g_strErrorDialogMsg_VerificationFailed              = self.m_dictLanguages[g_strDefaultLang]["STR_ERRORDIALOG_MSG_VERIFICATIONFAILED"]
        self.btnManualVerification.text                     = self.m_dictLanguages[g_strDefaultLang]["STR_BTN_MANUALVERIFICATION_TEXT"]
        self.lblMigration_boxSettingsBody.text              = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_MIGRATION_TEXT"]
        self.btnMigrateFromAnki_boxSettingsBody.text        = self.m_dictLanguages[g_strDefaultLang]["STR_BTN_MIGRATEFROMANKI_TEXT"]
        self.lblApkgFile_boxMigration.text                  = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_APKG_TEXT"]
        self.lblSelectWordList_boxMigration.text            = self.m_dictLanguages[g_strDefaultLang]["STR_LBL_SELECTWORDLIST_TEXT"]
        self.btnMigrate_boxMigration.text                   = self.m_dictLanguages[g_strDefaultLang]["STR_BTN_MIGRATE_TEXT"]
        global g_strInfoDialogTitle_Attention, g_strInfoDialogMsg_DoNotClose, g_strErrorDialogMsg_CanNotMigrate,\
            g_strLblMigrationInfo_Dyn_Step1, g_strLblMigrationInfo_Dyn_Step2, g_strLblMigrationInfo_Dyn_Step3, g_strLblMigrationInfo_Dyn_Step4,\
            g_strLblMigrationInfo_Dyn_Step5, g_strLblMigrationInfo_Dyn_Step6, g_strInfoDialogMsg_MigrationCompleted, g_strErrorDialogMsg_MigrationFailed
        g_strInfoDialogTitle_Attention                      = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_TITLE_ATTENTION']
        g_strInfoDialogMsg_DoNotClose                       = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_MSG_DONOTCLOSE']
        g_strErrorDialogMsg_CanNotMigrate                   = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_CANNOTMIGRATE']
        g_strLblMigrationInfo_Dyn_Step1                     = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_MIGRATIONINFO_DYN_STEP1']
        g_strLblMigrationInfo_Dyn_Step2                     = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_MIGRATIONINFO_DYN_STEP2']
        g_strLblMigrationInfo_Dyn_Step3                     = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_MIGRATIONINFO_DYN_STEP3']
        g_strLblMigrationInfo_Dyn_Step4                     = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_MIGRATIONINFO_DYN_STEP4']
        g_strLblMigrationInfo_Dyn_Step5                     = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_MIGRATIONINFO_DYN_STEP5']
        g_strLblMigrationInfo_Dyn_Step6                     = self.m_dictLanguages[g_strDefaultLang]['STR_LBL_MIGRATIONINFO_DYN_STEP6']
        g_strInfoDialogMsg_MigrationCompleted               = self.m_dictLanguages[g_strDefaultLang]['STR_INFODIALOG_MSG_MIGRATIONCOMPLETED']
        g_strErrorDialogMsg_MigrationFailed                 = self.m_dictLanguages[g_strDefaultLang]['STR_ERRORDIALOG_MSG_MIGRATIONFAILED']

    def ChangeSyncBtnText(self, iState): # WITHOUT COPY, it will cause abnormal phenomenon
        if iState == 0:
            self.btnSync_boxIndexBody.style.update(background_color = TRANSPARENT)
            self.btnSync_boxIndexBody.text = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_BTN_SYNC_DYN_SYNC"])
        elif iState == -1:
            self.btnSync_boxIndexBody.style.update(background_color = TRANSPARENT)
            self.btnSync_boxIndexBody.text = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_BTN_SYNC_DYN_SYNCING"])
        elif iState == 1:
            self.btnSync_boxIndexBody.style.update(background_color = "green")
            self.btnSync_boxIndexBody.text = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_BTN_SYNC_DYN_SYNCSUCCESS"])
        elif iState == 2:
            self.btnSync_boxIndexBody.style.update(background_color = "red")
            self.btnSync_boxIndexBody.text = "".join(self.m_dictLanguages[g_strDefaultLang]["STR_BTN_SYNC_DYN_SYNCFAILURE"])
        
    def FreshCards(self):
        while len(self.boxCards_boxLibraryBody.children) != 0:
            self.boxCards_boxLibraryBody.remove(self.boxCards_boxLibraryBody.children[0])
        for eachCard in self.m_dictTmpCards:
            boxCard = toga.Box(style = Pack(direction = ROW))
            mtxtCardInfo = toga.MultilineTextInput(style = Pack(flex = 1), readonly = True)
            strCardInfo = eachCard +"\n"
            if len(self.m_dictTmpCards[eachCard].keys()) == 1:
                strCardInfo += f"\t{g_strOnesided}\n"
                strCardInfo += f"\t\t{g_strFrontend} {g_strFilePath}:"+ self.m_dictTmpCards[eachCard]['single-sided']['frontend'] +"\n"
                strCardInfo += f"\t\t{g_strBackend} {g_strFilePath}:" + self.m_dictTmpCards[eachCard]['single-sided']['backend']
            else:
                strCardInfo += f"\t{g_strFront}\n"
                strCardInfo += f"\t\t{g_strFrontend} {g_strFilePath}:"+ self.m_dictTmpCards[eachCard]['front']['frontend'] +"\n"
                strCardInfo += f"\t\t{g_strBackend} {g_strFilePath}:" + self.m_dictTmpCards[eachCard]['front']['backend'] +"\n"
                
                strCardInfo += f"\t{g_strBack}\n"
                strCardInfo += f"\t\t{g_strFrontend} {g_strFilePath}:"+ self.m_dictTmpCards[eachCard]['back']['frontend'] +"\n"
                strCardInfo += f"\t\t{g_strBackend} {g_strFilePath}:" + self.m_dictTmpCards[eachCard]['back']['backend']
            mtxtCardInfo.value = strCardInfo
            
            boxCardOperation = toga.Box(style = Pack(direction = COLUMN))
            btnEditCard = toga.Button("🔧", on_press = partial(self.cbEditExistedCard, eachCard))
            btnDeleteCard = toga.Button("❌", on_press = partial(self.cbDeleteCard, eachCard))
            boxCardOperation.add(
                btnEditCard,
                btnDeleteCard
            )
            
            boxCard.add(
                mtxtCardInfo,
                boxCardOperation
            )
            self.boxCards_boxLibraryBody.add(boxCard)
    
    def FreshWordLists(self):
        while len(self.boxWordLists_boxLibraryBody.children) != 0:
            self.boxWordLists_boxLibraryBody.remove(self.boxWordLists_boxLibraryBody.children[0])
        lsAllCardTypes = [""]
        for eachCard in self.m_dictTmpCards:
            lsAllCardTypes.append(eachCard)
        for eachWordList in self.m_dictTmpWordLists:
            boxWordList = toga.Box(style = Pack(direction = COLUMN))

            boxWordListInfo = toga.Box(style = Pack(direction = ROW))
            txtWordListInfo = toga.TextInput(style = Pack(flex = 1), readonly = True)

            strWordListInfo = eachWordList + f" {g_strFilePath}:" + self.m_dictTmpWordLists[eachWordList]['FilePath']
            txtWordListInfo.value = strWordListInfo
            btnEditWordList = toga.Button("🔧", on_press = partial(self.cbEditExistedWordList, eachWordList))
            btnDeleteWordList = toga.Button("❌", on_press = self.cbDeleteWordList, id = eachWordList)
            boxWordListInfo.add(
                txtWordListInfo,
                btnEditWordList,
                btnDeleteWordList
            )

            boxWordListConfig = toga.Box(style = Pack(direction = COLUMN))

            boxWordListCards = toga.Box(style = Pack(direction = ROW, alignment = CENTER))

            lblCard1 = toga.Label(g_strCard1)
            cboCard1 = toga.Selection(style = Pack(flex = 1), items = lsAllCardTypes)
            cboCard1.value = self.m_dictTmpWordLists[eachWordList]["CardType1"]
            cboCard1.on_select = partial(self.cbChangeWordListCardType, eachWordList, 1)

            lblCard2 = toga.Label(g_strCard2)
            cboCard2 = toga.Selection(style = Pack(flex = 1), items = lsAllCardTypes)
            cboCard2.value = self.m_dictTmpWordLists[eachWordList]["CardType2"]
            cboCard2.on_select = partial(self.cbChangeWordListCardType, eachWordList, 2)

            boxWordListCards.add(
                lblCard1,
                cboCard1,
                lblCard2,
                cboCard2
            )

            boxWordListLearningPlan = toga.Box(style = Pack(direction = ROW, alignment = CENTER))

            numNewWordsPerGroup = toga.NumberInput(style = Pack(flex = 1, text_align = RIGHT), min_value = 1, \
                                                   on_change = partial(self.cbChangeNewWordsPerGroup, eachWordList))

            numNewWordsPerGroup.value = self.m_dictTmpWordLists[eachWordList]["NewWordsPerGroup"]

            lblNewWordsPerGroup = toga.Label(g_strLblNewWordsPerGroup)

            numOldWordsPerGroup = toga.NumberInput(style = Pack(flex = 1, padding_left = 5, text_align = RIGHT), min_value = 1,\
                                                   on_change = partial(self.cbChangeOldWordsPerGroup, eachWordList))
            numOldWordsPerGroup.value = self.m_dictTmpWordLists[eachWordList]["OldWordsPerGroup"]

            lblOldWordsPerGroup = toga.Label(g_strLblOldWordsPerGroup)

            lblReviewPolicy = toga.Label(g_strLblReviewPolicy, style = Pack(padding_left = 5))

            '''# This is not supported
            cboReviewPolicy = toga.Selection(style = Pack(flex = 1),\
                                                items = [\
                                                    {"name": g_strCboReviewPolicy_item_LearnFirst,  "val": 0},\
                                                    {"name": g_strCboReviewPolicy_item_Random,      "val": 1},\
                                                    {"name": g_strCboReviewPolicy_item_ReviewFirst, "val": 2},\
                                                ],\
                                                accessor = "name",\
                                                on_select = partial(self.cbChangeReviewPolicy, eachWordList)\
                                            )
            '''
            cboReviewPolicy = toga.Selection(style = Pack(flex = 1),\
                                                items = [g_strCboReviewPolicy_item_LearnFirst, g_strCboReviewPolicy_item_Random, g_strCboReviewPolicy_item_ReviewFirst],\
                                                on_select = partial(self.cbChangeReviewPolicy, eachWordList)\
                                            )
            cboReviewPolicy.value = cboReviewPolicy.items[self.m_dictTmpWordLists[eachWordList]["policy"]]
            boxWordListLearningPlan.add(
                numNewWordsPerGroup,
                lblNewWordsPerGroup,
                numOldWordsPerGroup,
                lblOldWordsPerGroup,
                lblReviewPolicy,
                cboReviewPolicy
            )
            boxWordListConfig.add(
                boxWordListCards,
                boxWordListLearningPlan
            )

            boxWordList.add(
                boxWordListInfo, 
                boxWordListConfig
            )
            self.boxWordLists_boxLibraryBody.add(boxWordList)


    def BackToLibraryAndFresh(self):
        while len(self.boxBody.children) != 0:
            self.boxBody.remove(self.boxBody.children[0])
        self.btnIndex_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.btnSettings_boxNavigateBar.style.update(font_weight = NORMAL, background_color = g_clrBg, color = 'black')
        self.btnLibrary_boxNavigateBar.style.update(font_weight = BOLD, background_color = g_clrPressedNavigationBtn, color = '#1651AA')
        self.boxBody.add(
            self.boxLibraryBody
        )
        self.FreshCards()
        self.FreshWordLists()
    
    def ChangeWordLearningBtns(self, iStatus):
        if iStatus == 0: # Clear all buttons
            self.boxIndexBody_Row3Hidden.style.update(visibility = HIDDEN)
            while len(self.boxIndexBody_Row3Blank.children) != 0:
                self.boxIndexBody_Row3Blank.remove(self.boxIndexBody_Row3Blank.children[0])
            while len(self.boxIndexBody_Row5.children) != 0:
                self.boxIndexBody_Row5.remove(self.boxIndexBody_Row5.children[0])
        elif iStatus == 1: # Show the front
            self.boxIndexBody_Row5.add(
                self.btnShowBack
            )
        elif iStatus == 2: # Show the back
            self.boxIndexBody_Row3Hidden.style.update(visibility = VISIBLE)
            self.boxIndexBody_Row3Blank.add(
                self.sldDelete
            )#self.btnDelete
            self.boxIndexBody_Row5.add(
                self.btnStrange,
                self.btnVague,
                self.btnFamiliar
            )

    def GenerateNewWordsSeqOfCurrentWordList(self):
        conn = sqlite3.connect(g_strDBPath)
        cursor = conn.cursor()
        cursor.execute("SELECT word_no FROM statistics WHERE wordlist = ? AND review_date IS NULL AND card_type = 1",(g_strOpenedWordListName,))
        lsCard_Type1 = [elem[0] for elem in cursor.fetchall()]
        cursor.execute("SELECT word_no FROM statistics WHERE wordlist = ? AND review_date IS NULL AND card_type = 2",(g_strOpenedWordListName,))
        lsCard_Type2 = [elem[0] for elem in cursor.fetchall()]
        conn.commit()
        conn.close()
        
        self.m_lsRememberSeq = []

        # No new words
        if len(lsCard_Type1) == len(lsCard_Type2) == 0:
            return
        
        # Only one card type
        if len(lsCard_Type2) == 0:
            for i in range(len(lsCard_Type1)):
                self.m_lsRememberSeq.append((lsCard_Type1[i],1))

        # Rest type-1 cards are less than rest type-2 cards 
        if len(lsCard_Type1) < len(lsCard_Type2):
            if len(lsCard_Type1) == 0: # type-1 cards are all studied
                for i in range(len(lsCard_Type2)):
                    self.m_lsRememberSeq.append((lsCard_Type2[i],2))
                return
            else: #Otherwise, first remember the extra type-2 cards
                for i in range(len(lsCard_Type2)):
                    if lsCard_Type2[i] != lsCard_Type1[0]:
                        self.m_lsRememberSeq.append((lsCard_Type2[i],2))
                    else:
                        break
        iNewWordsPerGroup = g_dictWordLists[g_strOpenedWordListName]['NewWordsPerGroup']
        lsGroups1 = []
        t = 1
        lsNumbersAGroup = []
        for i in range(len(lsCard_Type1)):
            if t == iNewWordsPerGroup or i == len(lsCard_Type1) - 1:
                lsNumbersAGroup.append([lsCard_Type1[i],1])
                lsGroups1.append(lsNumbersAGroup[:])
                lsNumbersAGroup.clear()
                t = 1
            else:
                lsNumbersAGroup.append([lsCard_Type1[i],1])
                t += 1
        lsGroups2 = copy.deepcopy(lsGroups1)
        for i in range(len(lsGroups2)):
            for j in range(len(lsGroups2[i])):
                lsGroups2[i][j][1] = 2
        result = [item for pair in zip(lsGroups1, lsGroups2) for item in pair]
        for elem in result:
            self.m_lsRememberSeq.extend(elem)

    def SaveWordListsAndCardsToFiles(self):
        global g_dictWordLists, g_dictCards
        jsonWordLists = json.dumps(g_dictWordLists)
        jsonCards = json.dumps(g_dictCards)
        with open(g_strWordlistsPath, 'w') as f:
            f.write(jsonWordLists)
        with open(g_strCardsPath, 'w') as f:
            f.write(jsonCards)

    def SaveConfiguration(self):
        (iX, iY) = self.main_window.position
        (iWidth, iHeight) = self.main_window.size
        json_str = json.dumps({
            "lang":         g_strDefaultLang,
            "cur_wordlist": g_strOpenedWordListName,
            "only_review":  g_bOnlyReview,
            "X":            iX,
            "Y":            iY,
            "width":        iWidth,
            "height":       iHeight,
            "appid":        g_strAppId, 
            "access_token": g_strAccessToken, 
            "username":     g_strUsername
        })
        if os.path.exists(g_strDataPath) == False:
            os.makedirs(g_strDataPath)    #In case of nonexistent directory
        with open(g_strConfigPath, 'w') as f:
            f.write(json_str)

    def ValidateFiles(self):
        if len(g_dictCards) == 0 or len(g_dictWordLists) == 0:
            return False
        try:
            # Validate cards
            for eachCard in g_dictCards:
                if len(g_dictCards[eachCard]) == 1:
                    strHtmlPath = g_dictCards[eachCard]['single-sided']['frontend']
                    with open(strHtmlPath, 'r') as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                    strPyPath = g_dictCards[eachCard]['single-sided']['backend']
                    with open(strPyPath, 'r') as f:
                        strCode = f.read()
                        compile(strCode, strPyPath, 'exec')
                else:

                    strFHtmlPath = g_dictCards[eachCard]['front']['frontend']
                    with open(strFHtmlPath, 'r') as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                    strFPyPath = g_dictCards[eachCard]['front']['backend']
                    with open(strFPyPath, 'r') as f:
                        strCode = f.read()
                        compile(strCode, strFPyPath, 'exec')

                    strBHtmlPath = g_dictCards[eachCard]['back']['frontend']
                    with open(strBHtmlPath, 'r') as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                    strBPyPath = g_dictCards[eachCard]['back']['backend']
                    with open(strBPyPath, 'r') as f:
                        strCode = f.read()
                        compile(strCode, strBPyPath, 'exec')

            for eachWordlist in g_dictWordLists:  
                # Validate wordlists
                strXlsxPath = g_dictWordLists[eachWordlist]['FilePath'] 
                xlsx = load_workbook(strXlsxPath, read_only = True)
                sheet = xlsx.worksheets[0]
                iMaxRowInSheet = sheet.max_row
                xlsx.close()
                
                conn = sqlite3.connect(g_strDBPath)
                cursor = conn.cursor()
                cursor.execute(f"SELECT MAX(word_no) FROM statistics WHERE wordlist=?",(eachWordlist,)) 
                iWMaxRowInDB = cursor.fetchone()[0]
                conn.commit()
                conn.close()
                if iMaxRowInSheet != iWMaxRowInDB:
                    raise FileNotFoundError
            return True
        except Exception:
            return False
        
def main():
    return wordgets()
